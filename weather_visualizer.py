# -*- coding: utf-8 -*-
"""
只見川流域の気象データ（気象庁ダウンロード形式のExcel / CSV）を自動で読み込み、
グラフ化・霧予測をするプログラム。

【読み込む気象データ】
  日時 / 気温(℃) / 降水量(mm) / 風速(m/s) / 風向 / 露点温度(℃) / 相対湿度(％)
  → 既定では見出し行の文字から【自動検出】します（--layout auto）。
    自動検出に失敗する場合は --layout fixed で従来どおりの固定列
    （A=日時, B=気温, E=降水量, H=風速, S=露点温度, V=相対湿度, AC〜BH=現象コード）を使えます。
  → 見出し行の【下】に続く小見出し行（「風向」「品質情報」など）も見出しとして読み、
    データ行としては読み飛ばします。

【現象コード（地点ごとに1列）】
  "/"      = その時刻に現象なし（コード0として扱う）
  空白     = まだデータ未入力（欠測。学習からは除外される）
  1〜10    = 現象発生コード
      1  薄い川霧    2  川霧      3  濃い川霧
      4  薄い全体霧  5  全体霧    6  全体濃い霧
      7  薄い層雲    8  濃い層雲  9  霧雨        10 雨
  → 地点の列かどうかは見出しと中身の両方で確かめます。水温のように「/」でも1〜10でもない
    値が並ぶ列は地点とみなさず、除外した理由を実行時に表示します。

【予測モデル】
  地点ごとに RandomForest と XGBoost を学習し、予測確率を加重平均するハイブリッド。
  混合比は地点ごとに検証期間（学習データの後ろ20%）の霧F1が最大になるよう自動決定します
  （比率0または1＝片方単独も候補）。--model rf / --model xgb で片方だけにもできます。
  xgboost はColabに標準搭載。無い環境では自動的にRandomForest単独へ切り替わります。

【出力】
  ①② 月別の気象データ×現象コード（全地点レーン）
  ④   地点ごとの予測グラフ（実測＋今後の予報＋霧確率＋予測現象コード）
  ⑤   日別霧予測サマリー（全地点まとめ）
  ⑥   予測結果CSV（地点ごとの予測コードと霧確率）

使い方（Google Colab を想定しています）:

    # ① このファイルを置いたフォルダで
    from weather_visualizer import run
    run("/content/drive/MyDrive/只見_気象データ.xlsx")

    # ② コマンドとして実行する場合
    !python weather_visualizer.py 入力ファイル.xlsx [出力フォルダ] [オプション]
    !python weather_visualizer.py --help          # オプション一覧
    !python weather_visualizer.py --check-font    # 日本語の文字化け確認だけ行う

日本語フォントが無い環境（Colabの初期状態）では、初回実行時に
IPAゴシック → Noto Sans CJK → japanize-matplotlib の順に自動導入を試み、
最後に「本当に日本語が描けるか」をフォントのグリフまで確認します。
"""

from __future__ import annotations

import argparse
import glob
import json
import math
import os
import re
import shutil
import subprocess
import sys
import time
import unicodedata
from datetime import date, datetime, timedelta

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")  # 画面のないサーバー環境（Colab等）でも背景で安全にグラフを描画するための設定
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.ticker as mticker
import matplotlib.font_manager as fm
from matplotlib.patches import Patch
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

# --- 霧予測アドオン用の追加インポート ---
import requests
from sklearn.base import BaseEstimator, ClassifierMixin
from sklearn.model_selection import RandomizedSearchCV, TimeSeriesSplit
from sklearn.pipeline import Pipeline
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import accuracy_score, f1_score, make_scorer, precision_score, recall_score
from sklearn.utils.class_weight import compute_sample_weight

# XGBoost は Google Colab には標準で入っている。入っていない環境
# （素のローカル等）でもプログラム全体が動くよう、無ければ RandomForest 単独に
# 自動フォールバックする（学習時にその理由を表示する）。
try:
    from xgboost import XGBClassifier
    _HAS_XGBOOST = True
except Exception:
    XGBClassifier = None
    _HAS_XGBOOST = False

# joblib は scikit-learn の依存として必ず入っているので、追加ライブラリは増えない
from joblib import dump as _joblib_dump, load as _joblib_load


# ===========================================================================
# 共通ユーティリティ（表の桁揃え・進捗表示）
# ---------------------------------------------------------------------------
# 日本語（全角）と数字・記号（半角）が混在する表を、文字数ではなく「表示幅」で
# 揃えるためのヘルパー。f"{s:<16}" は文字数で詰めるため、全角の多い地点名
# （例:「伊奈川合流地点」）と少ない地点名（例:「木賊」）で表がずれてしまう。
# ===========================================================================

def _disp_width(s) -> int:
    """全角文字を2、半角を1として数えた表示幅（ターミナル・Colab出力の桁数に対応）。"""
    return sum(2 if unicodedata.east_asian_width(ch) in "FWA" else 1 for ch in str(s))


def _pad(s, width, align="<") -> str:
    """表示幅ベースで空白を埋める（`_disp_width` を使うため全角混在でも桁が揃う）。"""
    s = str(s)
    fill = " " * max(0, width - _disp_width(s))
    return (fill + s) if align == ">" else (s + fill)


def _progress_enabled() -> bool:
    """テストや非対話実行では進捗バーの上書き表示をしない（ログが読みにくくなるため）。"""
    return not bool(os.environ.get("WEATHER_VIZ_QUIET"))


class _ProgressBar:
    """時間のかかる繰り返し処理の進捗を1行に表示する（外部ライブラリを増やさない簡易版）。

    `\\r` で同じ行を上書きするため、地点ごとの学習やグラフ生成のような
    「無言の時間が続く」処理でも、今どこまで進んだかをその場で確認できる。
    """

    def __init__(self, total, prefix="", width=24, enabled=None):
        self.total = max(1, int(total))
        self.prefix = prefix
        self.width = width
        self.enabled = _progress_enabled() if enabled is None else enabled
        self._n = 0
        self._start = time.time()
        self._last_len = 0

    def update(self, n: int = 1, suffix: str = ""):
        self._n = min(self.total, self._n + n)
        if not self.enabled:
            return
        frac = self._n / self.total
        filled = int(self.width * frac)
        bar = "█" * filled + "░" * (self.width - filled)
        elapsed = time.time() - self._start
        line = f"  {self.prefix} [{bar}] {self._n}/{self.total}（{frac:.0%}） 経過{elapsed:.0f}秒 {suffix}"
        pad = max(0, self._last_len - len(line))
        sys.stdout.write("\r" + line + " " * pad)
        sys.stdout.flush()
        self._last_len = len(line)

    def close(self):
        if self.enabled:
            sys.stdout.write("\n")
            sys.stdout.flush()


# ===========================================================================
# 0. 日本語フォント自動設定・インストール機能（文字化け・豆腐文字対策）
# ===========================================================================

_CJK_FONT_KEYWORDS = [
    "noto sans cjk", "noto serif cjk", "noto sans jp", "notosansjp",
    "ipaex", "ipagothic", "ipa gothic", "ipapgothic", "ipamincho",
    "takao", "vl gothic", "vl pgothic", "yu gothic", "ms gothic", "ms mincho",
    "meiryo", "hiragino", "source han sans", "source han serif",
    "droid sans fallback", "kozuka", "sazanami", "migmix", "mplus", "m+",
    "unifont", "japanize",
]

# このプログラムのグラフに実際に出てくる日本語。フォントがこれらの字を持っているかで
# 「本当に日本語を描けるフォントか」を判定する（名前だけで判断すると豆腐文字になる）。
_JP_TEST_TEXT = "霧気温露点湿度風速地点予測"

# 探しにいくフォントの置き場所（Colab/Ubuntu、ユーザーローカル、macOS、Windows）
_FONT_SEARCH_DIRS = [
    "/usr/share/fonts", "/usr/local/share/fonts", os.path.expanduser("~/.fonts"),
    os.path.expanduser("~/.local/share/fonts"), "/Library/Fonts",
    os.path.expanduser("~/Library/Fonts"), "C:/Windows/Fonts",
]
_FONT_FILE_PATTERN = "**/*.[ot]t[fc]"


def _font_file_supports_japanese(path):
    """フォントファイルが日本語のグリフを持っているか調べる。

    戻り値: True=描ける / False=描けない（豆腐文字になる） / None=判定できなかった
    matplotlibはfontToolsに依存しているので、追加インストールなしで判定できる。
    """
    if not path or not os.path.isfile(str(path)):
        return None
    try:
        from fontTools.ttLib import TTFont
    except Exception:
        return None
    try:
        # .ttc/.otc は複数書体をまとめたファイルなので、どの書体かを指定して開く
        # （matplotlibのfindfontはフォント内の番号つきで返してくることがある。
        #   ただしパスがただの文字列の場合、str.index はメソッドなので採用しない）
        kwargs = {}
        if str(path).lower().endswith((".ttc", ".otc")):
            raw_index = getattr(path, "index", None)
            kwargs["fontNumber"] = raw_index if isinstance(raw_index, int) else 0
        with TTFont(str(path), lazy=True, **kwargs) as font:
            cmap = font.getBestCmap()
    except Exception:
        return None
    return all(ord(ch) in cmap for ch in _JP_TEST_TEXT)


def _is_placeholder_font(name, path):
    """『すべての文字を持っているが、実際には□などの代替表示しかできない』フォントか。

    matplotlibに同梱されている Last Resort は全コードポイントに字形を持つため、
    グリフの有無だけで判定すると日本語フォントとして選ばれてしまい、
    結局グラフが□だらけになる。こうしたフォントは候補から除外する。
    """
    lowered = name.lower()
    if any(k in lowered for k in ("last resort", "adobe blank", "noto sans symbols")):
        return True
    # matplotlib同梱のフォント（DejaVu・STIX・Computer Modern）に日本語は入っていない
    return "mpl-data" in str(path).replace("\\", "/")


def _find_cjk_font(deep_scan=True):
    """日本語が描けるフォントを探して (フォント名, ファイルパス) を返す。

    1) 名前が日本語フォントらしいものを優先し、実際に日本語グリフがあるか確認する
    2) 見つからなければ、登録されている全フォントのグリフを順に確認する
       （名前が独特な日本語フォントでも拾えるようにするため）
    """
    named, others = [], []
    for f in fm.fontManager.ttflist:
        entry = (f.name, f.fname)
        if _is_placeholder_font(f.name, f.fname):
            continue
        if any(k in f.name.lower() for k in _CJK_FONT_KEYWORDS):
            named.append(entry)
        else:
            others.append(entry)

    unverified = None
    for name, path in named:
        supported = _font_file_supports_japanese(path)
        if supported:
            return name, path
        if supported is None and unverified is None:
            unverified = (name, path)  # 判定できない場合は候補として保留

    if deep_scan:
        seen = set()
        for name, path in others:
            if path in seen:
                continue
            seen.add(path)
            if _font_file_supports_japanese(path):
                return name, path

    return unverified if unverified else (None, None)


def _register_font_files(verbose=False):
    """フォントの置き場所を走査して、matplotlibに未登録のフォントを登録する。"""
    known = {f.fname for f in fm.fontManager.ttflist}
    added = 0
    for directory in _FONT_SEARCH_DIRS:
        if not os.path.isdir(directory):
            continue
        for path in glob.glob(os.path.join(directory, _FONT_FILE_PATTERN), recursive=True):
            if path in known:
                continue
            try:
                fm.fontManager.addfont(path)
                added += 1
            except Exception:
                pass
    if added and verbose:
        print(f"  フォントを{added}件登録しました。")
    return added


def _rebuild_font_manager():
    """matplotlibのフォント一覧をキャッシュを使わずに作り直す（最後の手段）。"""
    try:
        fm.fontManager = fm._load_fontmanager(try_read_cache=False)
        return True
    except Exception:
        return False


def _run_command(cmd, timeout):
    """コマンドを実行し、(成功したか, 出力の末尾) を返す。"""
    try:
        result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                timeout=timeout, check=False, text=True, errors="replace")
        return result.returncode == 0, (result.stdout or "").strip()[-400:]
    except FileNotFoundError:
        return False, f"{cmd[0]} が見つかりません"
    except subprocess.TimeoutExpired:
        return False, f"{' '.join(cmd[:3])} がタイムアウトしました"
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


def _install_japanese_font(verbose=True):
    """日本語フォントの導入を順番に試す（Colabのような環境向け）。

    軽くて速いものから試す:
      1. apt-get install fonts-ipafont-gothic （数MB・数十秒）
      2. apt-get install fonts-noto-cjk       （大きいが確実）
      3. pip install japanize-matplotlib      （aptが使えない環境向け）
    """
    attempts = []
    if sys.platform.startswith("linux"):
        attempts += [
            (["apt-get", "install", "-y", "fonts-ipafont-gothic"], 300, "IPAゴシック"),
            (["apt-get", "install", "-y", "fonts-noto-cjk"], 600, "Noto Sans CJK"),
        ]
    attempts.append(([sys.executable, "-m", "pip", "install", "-q", "japanize-matplotlib"],
                     300, "japanize-matplotlib"))

    apt_updated = False
    for cmd, timeout, label in attempts:
        if verbose:
            print(f"  {label} の導入を試みます…（数十秒かかることがあります）")
            print(f"    $ {' '.join(cmd)}")
        ok, output = _run_command(cmd, timeout)
        if not ok and cmd[0] == "apt-get" and not apt_updated:
            # パッケージ一覧が古いと install が失敗するので、一度だけ update してやり直す
            if verbose:
                print("  パッケージ一覧を更新しています… （apt-get update）")
            _run_command(["apt-get", "update"], 300)
            apt_updated = True
            ok, output = _run_command(cmd, timeout)

        if ok:
            if cmd[0] != "apt-get":
                # japanize-matplotlibはimportした時点でフォントを登録してくれる
                try:
                    import japanize_matplotlib  # noqa: F401
                except Exception as e:
                    if verbose:
                        print(f"  japanize-matplotlib の読み込みに失敗しました: {e}")
            _register_font_files(verbose=verbose)
            name, path = _find_cjk_font()
            if name is None:
                _rebuild_font_manager()
                name, path = _find_cjk_font()
            if name:
                return name, path
            if verbose:
                print(f"  {label} を入れましたが、日本語フォントとして認識できませんでした。")
        elif verbose:
            print(f"  {label} の導入に失敗しました: {output.splitlines()[-1] if output else '原因不明'}")

    return None, None


def _apply_font(font_name):
    """指定したフォントを最優先にしつつ、既存のフォントを予備として残す。

    font.family に1つだけ指定すると、そのフォントに無い記号（℃ や ⚠ など）が
    豆腐文字になる。sans-serif の先頭に入れることで、無い字は他のフォントで
    補われるようにする。
    """
    sans = [f for f in plt.rcParams["font.sans-serif"] if f != font_name]
    plt.rcParams["font.sans-serif"] = [font_name] + sans
    plt.rcParams["font.family"] = "sans-serif"
    plt.rcParams["axes.unicode_minus"] = False


def verify_japanese_font():
    """いま matplotlib が実際に使うフォントを調べ、日本語が描けるかを返す。

    戻り値: (描けるか(True/False/None), フォント名, ファイルパス)
    rcParamsにフォント名を設定できても、matplotlibが解決に失敗して
    DejaVu Sans に落ちていることがある（この状態が豆腐文字の正体）。
    """
    try:
        prop = fm.FontProperties(family=plt.rcParams["font.family"])
        path = fm.findfont(prop, fallback_to_default=True)
    except Exception:
        return None, None, None
    try:
        name = fm.get_font(path).family_name
    except Exception:
        name = os.path.basename(path)
    return _font_file_supports_japanese(path), name, path


_FONT_HELP = """
  グラフの日本語が □（豆腐文字）になります。次のいずれかで解決できます。
    Google Colab / Ubuntu : !apt-get -y install fonts-ipafont-gothic
    どの環境でも          : pip install japanize-matplotlib
    フォントを直接指定    : --font "IPAexGothic"  （環境変数 WEATHER_VIZ_FONT でも可）
  インストール後は、Colabではランタイムを再起動せずそのまま再実行すれば反映されます。
"""


def setup_japanese_font(font_name=None, verbose: bool = True, allow_install=None):
    """グラフ用の日本語フォントを設定する。

    font_name を指定した場合はそれを使う（見つからなければ警告して自動選択に戻る）。
    見つからない場合は、環境に応じてフォントの導入まで試みる。
    最後に「本当に日本語が描けるか」を検証して結果を表示する。
    """
    if allow_install is None:
        allow_install = not os.environ.get("WEATHER_VIZ_NO_FONT_INSTALL")
    font_name = font_name or os.environ.get("WEATHER_VIZ_FONT")

    if font_name:
        available = {f.name for f in fm.fontManager.ttflist}
        if font_name not in available:
            _register_font_files()
            available = {f.name for f in fm.fontManager.ttflist}
        if font_name in available:
            _apply_font(font_name)
        elif verbose:
            print(f"【警告】指定されたフォント『{font_name}』が見つかりません。自動選択に切り替えます。")
            font_name = None

    if not font_name:
        name, _ = _find_cjk_font()
        if name is None:
            # 未登録のフォントファイルがあるかもしれないので、登録してから探し直す
            _register_font_files()
            name, _ = _find_cjk_font()
        if name is None and allow_install:
            if verbose:
                print("日本語フォントが見つからないため、自動インストールを試みます…")
            name, _ = _install_japanese_font(verbose=verbose)
        if name:
            _apply_font(name)
        else:
            plt.rcParams["axes.unicode_minus"] = False

    supported, used_name, used_path = verify_japanese_font()
    if verbose:
        if supported:
            print(f"日本語フォントを設定しました: {used_name}")
            print(f"  （フォントファイル: {used_path}）")
        elif supported is False:
            print(f"【警告】いま使われるフォント『{used_name}』は日本語のグリフを持っていません。")
            print(_FONT_HELP.rstrip())
        else:
            print(f"【注意】日本語フォントの確認ができませんでした（使用フォント: {used_name}）。")
            print("　グラフの文字が □ になる場合は、次を試してください。")
            print(_FONT_HELP.rstrip())
    return used_name if supported else None


setup_japanese_font(verbose=not bool(os.environ.get("WEATHER_VIZ_QUIET")))


# ===========================================================================
# 1. 列レイアウト設定と現象コード（色・ラベル）の定義
# ===========================================================================

# --- 従来の固定レイアウト（--layout fixed / 自動検出の最終フォールバックで使用） ---
COL_DATETIME = "A"
COL_TEMP = "B"
COL_PRECIP = "E"
COL_WIND = "H"
COL_DEWPOINT = "S"
COL_HUMID = "V"
COL_WIND_DIR = None  # 固定レイアウトでは風向の位置が不明なため未使用

PHENOMENA_RANGE = ("AC", "BH")
LOCATION_NAME_ROW = 3  # 固定レイアウトで地点名が書かれている行（1始まり）

MAIN_COLUMNS = {
    COL_TEMP: "気温(℃)",
    COL_PRECIP: "降水量(mm)",
    COL_WIND: "風速(m/s)",
    COL_DEWPOINT: "露点温度(℃)",
    COL_HUMID: "相対湿度(％)",
}

# 気象要素の見出しを自動検出するためのキーワード（左から順に探し、最初に一致した列を採用）
MEASURE_KEYWORDS = {
    "気温(℃)": ["気温"],
    "降水量(mm)": ["降水量", "降水"],
    "風速(m/s)": ["風速"],
    "露点温度(℃)": ["露点"],
    "相対湿度(％)": ["相対湿度", "湿度"],
}

# 必須ではないが、ファイルにあれば予測の特徴量として自動的に使う気象要素。
# 見つからなくても警告は出さない（`OPTIONAL_MEASURE_MIN_COVERAGE` 未満のときも同様、
# 特徴量として使わないだけで実行は止めない）。
OPTIONAL_MEASURE_KEYWORDS = {
    "日照時間(時間)": ["日照"],
    "気圧(hPa)": ["気圧", "現地気圧", "海面気圧"],
    "日射量(MJ/m2)": ["日射量", "日射"],
    "視程(km)": ["視程"],
}
# ファイル中の有効値の割合がこれ未満の任意項目は、学習データを大きく減らさないよう
# 特徴量に加えない（例: 値が数%しか入っていない列を使うと dropna で大半の行が消える）。
OPTIONAL_MEASURE_MIN_COVERAGE = 0.05

WIND_DIR_LABEL = "風向"
WIND_DIR_KEYWORDS = ["風向"]
DATETIME_KEYWORDS = ["年月日時", "年月日", "日時", "日付", "時刻"]

# 気象庁データや観測記録に含まれる補助列（これらは観測地点ではないので現象コード列とみなさない）
# 「視程」「日照」「気圧」「日射」は OPTIONAL_MEASURE_KEYWORDS で気象要素として扱うため、
# ここには含めない（含めると測定値の列自体が最初の判定で弾かれてしまう）。
NON_LOCATION_KEYWORDS = [
    "品質情報", "均質番号", "現象なし情報", "備考", "合計", "平均", "最大", "最小",
    "天気", "雲量", "積雪", "降雪", "蒸気圧", "番号", "単位",
    "水温", "水位", "流量", "流速", "波高", "潮位",
]

# 「〜(℃)」「〜(mm)」のように単位が付いた見出しは観測値の列であって地点名ではない。
# 地点名にも括弧が使われる（例:「(川口橋)」）ため、中身が単位のときだけ弾く。
MEASURE_UNIT_PATTERN = re.compile(
    r"[（(]\s*(℃|°C|度|℉|mm|cm|m/s|m/秒|%|％|km|hPa|MJ/m2|MJ/㎡|W/m2|時間|分|秒)\s*[）)]"
)

# 現象コード列かどうかを中身で確かめるときの許容範囲。
# 「/」でも 0〜10 の整数でもない値がこの割合を超えたら、地点の列ではないと判断する
# （例: 水温の列は 8.2 や 15.7 といった小数が並ぶ。整数部分だけ見ると 8＝濃い層雲 として
#  学習されてしまうため、必ず中身まで確認する）。
PHENOM_INVALID_TOLERANCE = 0.1
PHENOM_CHECK_MAX_VALUES = 3000  # 判定に使う値の上限（大きなファイルでも速く終わらせるため）

MAX_HEADER_SEARCH_ROWS = 15
MAX_SUBHEADER_ROWS = 3  # 見出し行の下に続く小見出し行（風向・品質情報など）の最大数

PHENOM_LABELS = {
    1: "薄い川霧", 2: "川霧", 3: "濃い川霧", 4: "薄い全体霧", 5: "全体霧",
    6: "全体濃い霧", 7: "薄い層雲", 8: "濃い層雲", 9: "霧雨", 10: "雨",
}

PHENOM_COLORS = {
    1: "#aed6f1", 2: "#3498db", 3: "#1a5276", 4: "#dcdde1", 5: "#909497",
    6: "#2c3e50", 7: "#f8c471", 8: "#d35400", 9: "#a9dfbf", 10: "#196f3d",
}

SLASH_COLOR = "#f9e79f"

# 16方位（気象庁表記）→ 風が吹いてくる方角（度）。Open-Meteoのwind_direction_10mと同じ定義。
WIND_DIR_DEGREES = {
    "北": 0.0, "北北東": 22.5, "北東": 45.0, "東北東": 67.5,
    "東": 90.0, "東南東": 112.5, "南東": 135.0, "南南東": 157.5,
    "南": 180.0, "南南西": 202.5, "南西": 225.0, "西南西": 247.5,
    "西": 270.0, "西北西": 292.5, "北西": 315.0, "北北西": 337.5,
}
CALM_WORDS = {"静穏", "静穏(0.0)", "calm", "×", "-"}


# ===========================================================================
# 2. セル値の正規化・日時パース・現象コードの解釈
# ===========================================================================

def _cell_text(value) -> str:
    """セルの値を比較しやすい文字列に正規化する（全角→半角、前後空白と改行の除去）。"""
    if value is None:
        return ""
    if isinstance(value, float) and np.isnan(value):
        return ""
    s = str(value)
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("　", " ").replace("\n", "").replace("\r", "")
    return s.strip()


# 「2024年6月28日24時」「2024/6/28 1:00」「6月28日1時」などを解釈するための正規表現
_DATETIME_RE = re.compile(
    r"^(?:(?P<year>\d{4})\s*[年/\-\.]\s*)?"
    r"(?P<month>\d{1,2})\s*[月/\-\.]\s*"
    r"(?P<day>\d{1,2})\s*日?"
    r"(?:[\sT]*(?P<hour>\d{1,2})\s*(?:時|:)\s*(?P<minute>\d{1,2})?\s*分?(?::\d{1,2})?)?$"
)

# Excelのシリアル値（1900年日付システム）とみなす範囲。1970年〜2100年あたり。
_EXCEL_SERIAL_MIN = 25000
_EXCEL_SERIAL_MAX = 80000
_EXCEL_EPOCH = datetime(1899, 12, 30)


def parse_datetime_value(value, prev: pd.Timestamp | None = None,
                         fallback_year: int | None = None) -> pd.Timestamp | None:
    """1セル分の日時を解釈する。

    気象庁形式で実際に出てくる次のケースに対応する:
      * datetime / date オブジェクト（Excelが日付として保持している場合）
      * Excelのシリアル値（数値のまま保存されている場合）
      * 「2024年6月28日1時」のような日本語表記
      * 「24時」＝翌日0時（気象庁は1時〜24時で1日を表す）
      * 年が省略された「6月28日1時」（直前の行の年を引き継ぎ、年をまたいだら+1年する）

    prev には直前に解釈できた日時を渡す（年の省略と年またぎの判定に使う）。
    解釈できなければ None を返す（呼び出し側で件数を数えて警告する）。
    """
    if value is None:
        return None
    if isinstance(value, pd.Timestamp):
        return None if pd.isna(value) else value
    if isinstance(value, datetime):
        return pd.Timestamp(value)
    if isinstance(value, date):
        return pd.Timestamp(datetime(value.year, value.month, value.day))

    if isinstance(value, (int, float, np.integer, np.floating)):
        v = float(value)
        if np.isnan(v):
            return None
        if _EXCEL_SERIAL_MIN <= v <= _EXCEL_SERIAL_MAX:
            return pd.Timestamp(_EXCEL_EPOCH + timedelta(days=v))
        return None

    s = _cell_text(value)
    if not s:
        return None

    m = _DATETIME_RE.match(s)
    if m:
        year = int(m.group("year")) if m.group("year") else None
        month = int(m.group("month"))
        day = int(m.group("day"))
        hour = int(m.group("hour")) if m.group("hour") else 0
        minute = int(m.group("minute")) if m.group("minute") else 0
        if not (1 <= month <= 12 and 1 <= day <= 31 and 0 <= hour <= 24 and 0 <= minute <= 59):
            return None

        if year is None:
            # 年が省略されている場合、直前の行の年を引き継ぐ（12月→1月なら翌年）
            if prev is not None:
                year = prev.year
                if (month, day) < (prev.month, prev.day) and (prev.month - month) >= 6:
                    year += 1
            elif fallback_year is not None:
                year = fallback_year
            else:
                return None

        # 「24時」は翌日の0時を表す（気象庁の1〜24時表記）
        add_day = False
        if hour == 24:
            hour = 0
            add_day = True
        try:
            ts = pd.Timestamp(datetime(year, month, day, hour, minute))
        except ValueError:
            return None
        return ts + pd.Timedelta(days=1) if add_day else ts

    # 上記で解釈できない形式（ISO形式など）は pandas に任せる
    ts = pd.to_datetime(s, errors="coerce")
    if pd.isna(ts):
        return None
    return pd.Timestamp(ts)


def parse_datetime_series(values, fallback_year: int | None = None):
    """日時セルの並びをまとめて解釈する。

    戻り値: (pd.Series[datetime64], 解釈できなかった値のリスト)
    直前に解釈できた日時を引き継ぐため、年が省略された行があっても
    「実行した年」に化けることがない（従来ここで年が今年になる不具合があった）。
    """
    if fallback_year is None:
        # 先に解釈できる行から年を拾っておく（先頭行だけ年が省略されている場合の保険）
        for v in values:
            ts = parse_datetime_value(v)
            if ts is not None:
                fallback_year = ts.year
                break

    parsed, unparsed = [], []
    prev = None
    for v in values:
        ts = parse_datetime_value(v, prev=prev, fallback_year=fallback_year)
        if ts is not None:
            prev = ts
        else:
            text = _cell_text(v)
            if text:
                unparsed.append(text)
        parsed.append(ts)
    return pd.Series(parsed, dtype="datetime64[ns]"), unparsed


def encode_phenomena_cell(value):
    """セルの値を内部コードに変換する。

      "/"（現象なし）     -> 0.0
      1〜10 の数字        -> その値
      空欄・解釈不能      -> NaN（欠測扱い。学習からは除外される）

    全角数字「１０」や全角スラッシュ「／」も正しく扱う。
    1セルに複数コードが書かれている場合は、より大きいコードを採用する。
    """
    if value is None:
        return np.nan
    if isinstance(value, (int, float, np.integer, np.floating)) and not isinstance(value, bool):
        v = float(value)
        if np.isnan(v):
            return np.nan
        code = int(v)
        return float(code) if 0 <= code <= 10 else np.nan

    s = _cell_text(value)
    if not s:
        return np.nan
    if s == "/":
        return 0.0
    nums = re.findall(r"\d+", s)
    if nums:
        code = max(int(n) for n in nums)
        return float(code) if 0 <= code <= 10 else np.nan
    return np.nan


def encode_wind_direction(value):
    """風向（「南南西」「静穏」など）を度に変換する。

    静穏（無風）は方角が定義できないため NaN を返し、特徴量では sin=cos=0 として扱う。
    """
    if value is None:
        return np.nan
    if isinstance(value, (int, float, np.integer, np.floating)) and not isinstance(value, bool):
        v = float(value)
        if np.isnan(v) or not (0 <= v <= 360):
            return np.nan
        return v % 360.0
    s = _cell_text(value)
    if not s:
        return np.nan
    if s in CALM_WORDS:
        return np.nan
    if s in WIND_DIR_DEGREES:
        return WIND_DIR_DEGREES[s]
    try:
        v = float(s)
    except ValueError:
        return np.nan
    return v % 360.0 if 0 <= v <= 360 else np.nan


# ===========================================================================
# 3. データ読み込み（Excel / CSV）と列レイアウトの自動検出
# ===========================================================================

def _read_grid_from_excel(filepath, sheet_name=None):
    """Excelを2次元リスト（行×列のセル値）として読み込む。"""
    wb = load_workbook(filepath, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _read_grid_from_csv(filepath):
    """CSVを2次元リストとして読み込む（UTF-8 / BOM付き / Shift_JISに対応）。"""
    last_error = None
    for encoding in ("utf-8-sig", "cp932", "utf-8"):
        try:
            df = pd.read_csv(filepath, header=None, dtype=str, encoding=encoding,
                             keep_default_na=False, engine="python")
            break
        except (UnicodeDecodeError, pd.errors.ParserError) as e:  # 次のエンコーディングを試す
            last_error = e
    else:
        raise ValueError(f"CSVを読み込めませんでした: {last_error}")
    return [[(v if v != "" else None) for v in row] for row in df.itertuples(index=False, name=None)]


def read_grid(filepath, sheet_name=None):
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".csv":
        return _read_grid_from_csv(filepath)
    if ext in (".xlsx", ".xlsm", ".xls"):
        return _read_grid_from_excel(filepath, sheet_name)
    raise ValueError(f"対応していないファイル形式です: {ext}（.xlsx / .xlsm / .csv に対応）")


def _grid_get(grid, r, c):
    if 0 <= r < len(grid) and 0 <= c < len(grid[r]):
        return grid[r][c]
    return None


def _grid_width(grid):
    return max((len(row) for row in grid), default=0)


def find_header_row(grid, keyword="年月日時", max_search_rows=MAX_HEADER_SEARCH_ROWS):
    """見出し行（0始まりの行番号）を探す。

    1) A列が「年月日時」の行（気象庁ダウンロード形式）
    2) 気象要素のキーワードが2つ以上並んでいる行
    どちらも見つからなければ 0 行目を見出し行とみなす。
    """
    limit = min(max_search_rows, len(grid))
    for r in range(limit):
        if _cell_text(_grid_get(grid, r, 0)) == keyword:
            return r

    best_row, best_hits = 0, 0
    all_keywords = [k for keys in MEASURE_KEYWORDS.values() for k in keys]
    for r in range(limit):
        texts = [_cell_text(v) for v in grid[r]]
        hits = sum(1 for kw in all_keywords if any(kw in t for t in texts))
        if hits > best_hits:
            best_row, best_hits = r, hits
    if best_hits >= 2:
        return best_row
    return 0


def _header_texts(grid, col_idx, header_row, sub_rows=()):
    """その列の見出し（見出し行までの各行の文字＋小見出し行の文字）を上から並べて返す。

    気象庁形式では「風向」のように、見出し行の【下】の行に書かれている項目がある。
    sub_rows（`_find_subheader_rows()` の戻り値）を渡すとその行も見出しとして扱う。
    """
    rows = list(range(0, header_row + 1)) + list(sub_rows)
    return [t for r in rows if (t := _cell_text(_grid_get(grid, r, col_idx)))]


def _find_subheader_rows(grid, header_row, dt_col):
    """見出し行のすぐ下に続く「小見出し行」の行番号を返す。

    気象庁のExcelは見出し行の下に「風向」「品質情報」などの補助的な見出しが
    1〜2行続く。データ行には必ず日時が入っているので、日時の欄が空のあいだを
    小見出し行とみなす（これらは日時として解釈できない行として毎回除外されていた）。
    """
    rows = []
    for r in range(header_row + 1, min(header_row + 1 + MAX_SUBHEADER_ROWS, len(grid))):
        if _cell_text(_grid_get(grid, r, dt_col)):
            break
        rows.append(r)
    return rows


def _is_metadata_header(texts):
    return any(any(kw in t for kw in NON_LOCATION_KEYWORDS) for t in texts)


def _is_measure_header(texts):
    all_keys = ([k for keys in MEASURE_KEYWORDS.values() for k in keys]
                + [k for keys in OPTIONAL_MEASURE_KEYWORDS.values() for k in keys]
                + WIND_DIR_KEYWORDS + DATETIME_KEYWORDS)
    if any(any(kw in t for kw in all_keys) for t in texts):
        return True
    # 「水温(℃)」のように既知のキーワードに無い観測項目でも、単位が付いていれば観測値の列
    return any(MEASURE_UNIT_PATTERN.search(t) for t in texts)


def _is_phenomena_value(value):
    """そのセルの値が現象コードとして自然か（「/」または0〜10の整数か）を判定する。

    「1 8」のように1セルに複数コードが書かれる記法も現象コードとして扱う。
    小数（8.2 など）や11以上の数値は現象コードではない＝地点の列ではない目印になる。
    """
    if isinstance(value, (int, float, np.integer, np.floating)) and not isinstance(value, bool):
        v = float(value)
        return (not np.isnan(v)) and float(v).is_integer() and 0 <= v <= 10

    s = _cell_text(value)
    if not s:
        return False  # 空欄は呼び出し側で除外済み
    if s == "/":
        return True
    tokens = [t for t in re.split(r"[\s,、/／・]+", s) if t]
    if not tokens:
        return False
    return all(re.fullmatch(r"\d{1,2}", t) and int(t) <= 10 for t in tokens)


def _check_phenomena_column(grid, col_idx, data_start):
    """現象コードの列（＝観測地点の列）らしいかを、実際に入っている値から確かめる。

    戻り値: (地点の列とみなしてよいか, 現象コードではなかった値の例)
    値が1つも入っていない列は「これから記入する地点」の可能性があるので通す。
    """
    bad, n_values = [], 0
    for r in range(data_start, len(grid)):
        value = _grid_get(grid, r, col_idx)
        if value is None or _cell_text(value) == "":
            continue
        n_values += 1
        if not _is_phenomena_value(value):
            bad.append(_cell_text(value))
        if n_values >= PHENOM_CHECK_MAX_VALUES:
            break
    if n_values == 0:
        return True, []
    return len(bad) <= n_values * PHENOM_INVALID_TOLERANCE, list(dict.fromkeys(bad))[:5]


def detect_layout(grid, header_row, layout="auto"):
    """列レイアウト（日時列・気象要素列・現象コード列）を決める。

    layout="fixed" のときは従来どおりの固定列を使う。
    layout="auto"  のときは見出しの文字から検出し、見つからない要素だけ固定列で補う。

    戻り値: {"datetime": 列番号, "measures": {ラベル: 列番号|None},
             "wind_dir": 列番号|None, "phenom": [(列文字, 列番号, 地点名), ...],
             "excluded": [(列文字, 見出し, 除外理由), ...], "data_start": データ開始行,
             "source": "auto"|"fixed"}
    """
    width = _grid_width(grid)

    if layout == "fixed":
        return _fixed_layout(grid, width, header_row)

    dt_col = None
    for c in range(width):
        texts = _header_texts(grid, c, header_row)
        if texts and any(any(kw in t for kw in DATETIME_KEYWORDS) for t in texts):
            dt_col = c
            break
    if dt_col is None:
        dt_col = 0

    # 見出し行の下に続く小見出し行（「風向」など）も見出しとして読む
    sub_rows = _find_subheader_rows(grid, header_row, dt_col)
    data_start = header_row + 1 + len(sub_rows)

    all_measure_keywords = {**MEASURE_KEYWORDS, **OPTIONAL_MEASURE_KEYWORDS}
    measures = {label: None for label in all_measure_keywords}
    wind_dir_col = None
    for c in range(width):
        if c == dt_col:
            continue
        texts = _header_texts(grid, c, header_row, sub_rows)
        if not texts or _is_metadata_header(texts):
            continue
        for label, keywords in all_measure_keywords.items():
            if measures[label] is None and any(kw in t for kw in keywords for t in texts):
                measures[label] = c
                break
        else:
            if wind_dir_col is None and any(kw in t for kw in WIND_DIR_KEYWORDS for t in texts):
                wind_dir_col = c

    used = [c for c in measures.values() if c is not None]
    if wind_dir_col is not None:
        used.append(wind_dir_col)
    last_measure_col = max(used) if used else dt_col

    # 気象要素より右にあり、名前が付いていて、補助列でもない列を観測地点とみなす
    phenom = []
    excluded = []
    seen_names = {}
    for c in range(last_measure_col + 1, width):
        letter = get_column_letter(c + 1)
        texts = _header_texts(grid, c, header_row, sub_rows)
        if not texts or _is_metadata_header(texts) or _is_measure_header(texts):
            if texts:
                excluded.append((letter, texts[0], "観測値・補助情報の列（地点名ではない）"))
            continue
        name = texts[0]
        if re.fullmatch(r"[\d\.\-/]+", name):  # 数字だけの見出しは地点名ではない
            continue
        # 見出しだけでは判断できない列があるため、中身が現象コードかどうかも確かめる
        # （水温のように「地点名らしい見出し＋数値」の列を学習に混ぜないため）
        ok, bad_samples = _check_phenomena_column(grid, c, data_start)
        if not ok:
            reason = "現象コード（「/」または1〜10）以外の値が多い"
            if bad_samples:
                reason += f"（例: {', '.join(bad_samples)}）"
            excluded.append((letter, name, reason))
            continue
        if name in seen_names:  # 同名地点は列文字を付けて区別する
            name = f"{name}({letter})"
        seen_names[name] = c
        phenom.append((letter, c, name))

    if not phenom or all(m is None for m in measures.values()):
        # 自動検出がうまくいかなかった場合は従来の固定レイアウトに戻す
        fixed = _fixed_layout(grid, width, header_row)
        if not phenom:
            phenom = fixed["phenom"]
            excluded = fixed["excluded"]
        for label, col in measures.items():
            if col is None:
                measures[label] = fixed["measures"].get(label)
        return {"datetime": dt_col, "measures": measures, "wind_dir": wind_dir_col,
                "phenom": phenom, "excluded": excluded, "data_start": data_start,
                "source": "auto+fixed"}

    return {"datetime": dt_col, "measures": measures, "wind_dir": wind_dir_col,
            "phenom": phenom, "excluded": excluded, "data_start": data_start,
            "source": "auto"}


def _fixed_layout(grid, width, header_row=None):
    """従来のハードコードされた列レイアウト。"""
    measures = {}
    for letter, label in MAIN_COLUMNS.items():
        measures[label] = column_index_from_string(letter) - 1
    wind_dir_col = column_index_from_string(COL_WIND_DIR) - 1 if COL_WIND_DIR else None

    dt_col = column_index_from_string(COL_DATETIME) - 1
    if header_row is None:
        header_row = LOCATION_NAME_ROW - 1
    data_start = header_row + 1 + len(_find_subheader_rows(grid, header_row, dt_col))

    start_idx = column_index_from_string(PHENOMENA_RANGE[0]) - 1
    end_idx = column_index_from_string(PHENOMENA_RANGE[1]) - 1
    phenom, excluded = [], []
    for c in range(start_idx, min(end_idx, width - 1) + 1):
        raw = _cell_text(_grid_get(grid, LOCATION_NAME_ROW - 1, c))
        letter = get_column_letter(c + 1)
        name = raw if raw else f"地点_{letter}"
        ok, bad_samples = _check_phenomena_column(grid, c, data_start)
        if not ok:
            reason = "現象コード（「/」または1〜10）以外の値が多い"
            if bad_samples:
                reason += f"（例: {', '.join(bad_samples)}）"
            excluded.append((letter, name, reason))
            continue
        phenom.append((letter, c, name))
    return {"datetime": dt_col, "measures": measures, "wind_dir": wind_dir_col,
            "phenom": phenom, "excluded": excluded, "data_start": data_start,
            "source": "fixed"}


def load_weather_data(filepath, sheet_name=None, layout="auto", verbose=True):
    """Excel/CSVを読み込み、気象データ・現象コード・地点名の対応を作る。

    戻り値: (main_df, phenom_df, phenom_cols, location_mapping)
      main_df       : datetime + 気象要素（+ 風向）
      phenom_df     : datetime + 地点ごとの生セル値（列名は列文字）
      phenom_cols   : 地点列の列文字リスト（左から順）
      location_mapping : {列文字: 地点名}
    """
    grid = read_grid(filepath, sheet_name)
    if not grid:
        raise ValueError("ファイルが空です。")

    header_row = find_header_row(grid)
    info = detect_layout(grid, header_row, layout=layout)
    # 見出し行の下に小見出し行（風向・品質情報など）がある形式では、そこを飛ばして読む
    data_start = info.get("data_start", header_row + 1)
    data_rows = grid[data_start:]

    dt_col = info["datetime"]
    raw_dt = [_grid_get(grid, data_start + i, dt_col) for i in range(len(data_rows))]
    dt_series, unparsed = parse_datetime_series(raw_dt)

    main_data = {"datetime": dt_series}
    missing_measures = []
    for label in MAIN_COLUMNS.values():
        col = info["measures"].get(label)
        if col is None:
            missing_measures.append(label)
            main_data[label] = pd.Series([np.nan] * len(data_rows), dtype="float64")
        else:
            values = [_grid_get(grid, data_start + i, col) for i in range(len(data_rows))]
            main_data[label] = pd.to_numeric(pd.Series(values, dtype="object"), errors="coerce")

    # 任意項目（日照時間・気圧など）は見つかった場合だけ列を作る。
    # 見つからない列を毎回NaN列として増やすと、後続の処理が煩雑になるだけなので作らない。
    for label in OPTIONAL_MEASURE_KEYWORDS:
        col = info["measures"].get(label)
        if col is not None:
            values = [_grid_get(grid, data_start + i, col) for i in range(len(data_rows))]
            main_data[label] = pd.to_numeric(pd.Series(values, dtype="object"), errors="coerce")

    if info["wind_dir"] is not None:
        wd = [_grid_get(grid, data_start + i, info["wind_dir"]) for i in range(len(data_rows))]
        main_data[WIND_DIR_LABEL] = pd.Series([encode_wind_direction(v) for v in wd], dtype="float64")

    phenom_data = {"datetime": dt_series}
    phenom_cols, location_mapping = [], {}
    for letter, col, name in info["phenom"]:
        phenom_cols.append(letter)
        location_mapping[letter] = name
        phenom_data[letter] = pd.Series(
            [_grid_get(grid, data_start + i, col) for i in range(len(data_rows))], dtype="object")

    main_df = pd.DataFrame(main_data)
    phenom_df = pd.DataFrame(phenom_data)

    valid = main_df["datetime"].notna()
    n_dropped = int((~valid).sum())
    main_df = main_df[valid].sort_values("datetime").reset_index(drop=True)
    phenom_df = phenom_df[valid].sort_values("datetime").reset_index(drop=True)

    # 同じ日時の行が重複していると、後段のmergeで行数が増えてしまうため、後の行を残して除去する
    dup = main_df.duplicated("datetime", keep="last")
    n_dup = int(dup.sum())
    if n_dup:
        main_df = main_df[~dup].reset_index(drop=True)
        phenom_df = phenom_df[~phenom_df.duplicated("datetime", keep="last")].reset_index(drop=True)

    if main_df.empty:
        raise ValueError(
            "日時として解釈できる行が1件もありませんでした。"
            f"（見出し行={header_row + 1}行目, 日時列={get_column_letter(dt_col + 1)}列）"
        )
    if not phenom_cols:
        raise ValueError(
            "現象コードの列（観測地点）が1つも見つかりませんでした。"
            f"（列レイアウト={info['source']}）\n"
            "　・自動検出の場合: 地点名が気象要素の列より右の行に書かれているか確認してください。\n"
            "　・--layout fixed の場合: AC〜BH列に現象コードがあるか確認してください。"
        )

    if verbose:
        _print_load_report(filepath, header_row, info, main_df, phenom_cols, location_mapping,
                           unparsed, n_dropped, n_dup, missing_measures)

    return main_df, phenom_df, phenom_cols, location_mapping


def _print_load_report(filepath, header_row, info, main_df, phenom_cols, location_mapping,
                       unparsed, n_dropped, n_dup, missing_measures):
    """どのファイルのどの列から何を読んだのかを表示する（データの出どころの確認用）。"""
    print("\n" + "=" * 66)
    print("■ 【確認】読み込んだファイルと列の対応")
    print("=" * 66)
    print(f"  ファイル : {os.path.abspath(filepath)}")
    print(f"  見出し行 : {header_row + 1}行目　列レイアウト: {info['source']}")
    print(f"  日時列   : {get_column_letter(info['datetime'] + 1)}列")
    for label, col in info["measures"].items():
        is_optional = label in OPTIONAL_MEASURE_KEYWORDS
        if is_optional and col is None:
            continue  # 任意項目が無いのは普通のことなので、毎回は表示しない
        where = f"{get_column_letter(col + 1)}列" if col is not None else "見つからず（欠測扱い）"
        n_valid = int(main_df[label].notna().sum()) if label in main_df else 0
        coverage = n_valid / len(main_df) if len(main_df) else 0.0
        if is_optional:
            tag = ("（任意・特徴量として使用）" if coverage >= OPTIONAL_MEASURE_MIN_COVERAGE
                  else "（任意・有効値が少ないため特徴量には使いません）")
        else:
            tag = ""
        print(f"  {_pad(label, 14)}: {_pad(where, 20)} 有効値 {n_valid}件{tag}")
    if info["wind_dir"] is not None:
        print(f"  {_pad(WIND_DIR_LABEL, 14)}: {_pad(get_column_letter(info['wind_dir'] + 1) + '列', 20)}"
              f" 有効値 {int(main_df[WIND_DIR_LABEL].notna().sum())}件")
    print(f"  期間     : {main_df['datetime'].min()} 〜 {main_df['datetime'].max()}（{len(main_df)}行）")

    if missing_measures:
        print(f"  【注意】次の要素の列が見つかりませんでした: {', '.join(missing_measures)}")
    if n_dropped:
        print(f"  【注意】日時として解釈できず除外した行: {n_dropped}件")
        if unparsed:
            samples = ", ".join(dict.fromkeys(unparsed[:5]))
            print(f"           例: {samples}")
    if n_dup:
        print(f"  【注意】日時が重複していた行を{n_dup}件除外しました（後の行を採用）。")

    print(f"\n  検出した観測地点（{len(phenom_cols)}地点）:")
    for letter in phenom_cols:
        print(f"    {letter}列 → {location_mapping.get(letter, '(不明)')}")

    # 「地点ではない列（水温など）まで学習していないか」を毎回確認できるようにする
    excluded = info.get("excluded", [])
    if excluded:
        print(f"\n  地点として扱わなかった列（{len(excluded)}列）:")
        for letter, name, reason in excluded[:10]:
            print(f"    {letter}列 → {name}: {reason}")
        if len(excluded) > 10:
            print(f"    （ほか{len(excluded) - 10}列）")


def report_phenomena_quality(phenom_df, phenom_cols, location_mapping):
    """地点ごとに「/」（現象なし）・現象コード・未入力の件数を表示する。

    「/」が1件もない地点は、現象があった時刻だけを記入している可能性が高く、
    そのままでは『霧でない例』を学習できない。ここで気づけるようにしておく。
    """
    print("\n" + "=" * 66)
    print("■ 【確認】地点ごとの記録状況（学習に使えるデータ量）")
    print("=" * 66)
    print(f"  {'地点名':<16}{'記録あり':>8}{'うち「/」':>10}{'うち現象':>9}{'未入力':>8}  最終記録日")
    print("  " + "-" * 64)
    warnings = []
    for col in phenom_cols:
        codes = phenom_df[col].map(encode_phenomena_cell)
        recorded = codes.notna()
        n_rec = int(recorded.sum())
        n_slash = int((codes == 0).sum())
        n_event = n_rec - n_slash
        n_blank = int(len(codes) - n_rec)
        last = phenom_df.loc[recorded, "datetime"].max() if n_rec else None
        last_str = last.strftime("%Y/%m/%d %H時") if last is not None and pd.notna(last) else "-"
        name = location_mapping.get(col, col)
        print(f"  {name:<16}{n_rec:>8}{n_slash:>10}{n_event:>9}{n_blank:>8}  {last_str}")
        if n_rec and n_slash == 0:
            warnings.append(
                f"{name}: 「/」（現象なし）の記録が0件です。現象があった時刻だけ記入されている場合、"
                "『現象なし』の例が学習できず、予測が現象側に偏ります。")
    for w in warnings:
        print(f"  ⚠ {w}")


# ===========================================================================
# 4. グラフ生成・描画処理（上下2段構成・タイムライン同期システム）
# ===========================================================================

DAYS_PER_INCH = 0.55
MIN_FIG_WIDTH = 16
MAX_FIG_WIDTH = 60
SAVE_DPI = 150


def compute_fig_width(times):
    times = pd.Series(pd.to_datetime(pd.Series(times).dropna()))
    if times.empty:
        return float(MIN_FIG_WIDTH)
    total_days = (times.max() - times.min()).total_seconds() / 86400.0
    width = max(total_days, 1.0) / DAYS_PER_INCH
    return float(np.clip(width, MIN_FIG_WIDTH, MAX_FIG_WIDTH))


def split_by_month_two(main_df, phenom_df):
    m = main_df.copy()
    p = phenom_df.copy()
    m["__ym"] = m["datetime"].dt.strftime("%Y-%m")
    p["__ym"] = p["datetime"].dt.strftime("%Y-%m")
    groups = []
    for ym in sorted(set(m["__ym"]) | set(p["__ym"])):
        msub = m[m["__ym"] == ym].drop(columns="__ym").reset_index(drop=True)
        psub = p[p["__ym"] == ym].drop(columns="__ym").reset_index(drop=True)
        groups.append((ym, msub, psub))
    return groups


def _phenom_legend_handles():
    handles = [Patch(facecolor=SLASH_COLOR, label="「/」現象なし")]
    for code in range(1, 11):
        handles.append(Patch(facecolor=PHENOM_COLORS[code], label=f"{code}: {PHENOM_LABELS[code]}"))
    return handles


def _draw_lane_panel(ax2, phenom_df, phenom_cols, location_mapping):
    ptimes = phenom_df["datetime"]
    n_lanes = len(phenom_cols)

    for row, col_letter in enumerate(phenom_cols):
        col_values = phenom_df[col_letter].map(encode_phenomena_cell).to_numpy(dtype="float64")

        slash_mask = col_values == 0
        if slash_mask.any():
            x_vals = mdates.date2num(ptimes[slash_mask])
            ax2.vlines(x_vals, row - 0.35, row + 0.35,
                       color=SLASH_COLOR, linewidth=1.0, alpha=0.9, zorder=2)

        for code in range(1, 11):
            code_mask = col_values == code
            if code_mask.any():
                x_vals = mdates.date2num(ptimes[code_mask])
                ax2.vlines(x_vals, row - 0.42, row + 0.42,
                           color=PHENOM_COLORS[code], linewidth=2.4, zorder=3)

    for row in range(n_lanes + 1):
        ax2.axhline(row - 0.5, color="#ececec", linewidth=0.5, zorder=1)

    ax2.set_yticks(range(n_lanes))
    ax2.set_yticklabels([location_mapping[c] for c in phenom_cols], fontsize=8)
    ax2.set_ylim(n_lanes - 0.5, -0.5)
    ax2.set_xlabel("日時")
    ax2.set_ylabel("現象記録地点", fontsize=10)

    ax2.legend(handles=_phenom_legend_handles(), bbox_to_anchor=(1.0, 1.0), loc="upper left",
               fontsize=8, borderaxespad=0., title="現象コード", title_fontsize=9)


def _finalize_figure(fig, ax1, ax2, times, fig_width):
    ax2.xaxis_date()
    span_days = max(1.0, (times.max() - times.min()).total_seconds() / 86400.0)
    if span_days <= 3:
        # 数日分しかない場合は時刻まで表示する
        ax2.xaxis.set_major_locator(mdates.HourLocator(interval=3))
        ax2.xaxis.set_major_formatter(mdates.DateFormatter("%Y/%m/%d %H時"))
    else:
        # AutoDateLocatorに任せると1日に複数の目盛りが立ち、同じ日付ラベルが
        # 重複表示されてしまうため、1日単位の目盛りを明示的に指定する。
        ax2.xaxis.set_major_locator(mdates.DayLocator(interval=max(1, int(np.ceil(span_days / 40)))))
        ax2.xaxis.set_major_formatter(mdates.DateFormatter("%Y/%m/%d"))
    ax1.set_xlim(times.min(), times.max())
    fig.autofmt_xdate()


def _make_fig(times, phenom_cols):
    n_lanes = len(phenom_cols)
    fig_width = compute_fig_width(times)
    lane_panel_height = max(6.0, n_lanes * 0.24)
    fig_height = 5.5 + lane_panel_height
    fig, (ax1, ax2) = plt.subplots(
        2, 1, figsize=(fig_width, fig_height), sharex=True,
        gridspec_kw={"height_ratios": [5.5, lane_panel_height], "hspace": 0.04},
    )
    return fig, ax1, ax2, fig_width


def plot_temp_humid_dew(main_df, phenom_df, phenom_cols, location_mapping, location_name, out_path):
    times = main_df["datetime"]
    fig, ax1, ax2, fig_width = _make_fig(times, phenom_cols)

    l1, = ax1.plot(times, main_df["気温(℃)"],    color="#e74c3c", linewidth=1.1, label="気温(℃)", zorder=3)
    l2, = ax1.plot(times, main_df["露点温度(℃)"], color="#16a085", linewidth=1.1, label="露点温度(℃)", zorder=3)
    ax1.set_ylabel("気温・露点温度（℃）", fontsize=10)
    ax1.grid(True, alpha=0.25)

    ax1r = ax1.twinx()
    l3, = ax1r.plot(times, main_df["相対湿度(％)"], color="#8e44ad", linewidth=0.9, alpha=0.65,
                    label="相対湿度(％)", zorder=2)
    ax1r.set_ylabel("相対湿度（％）", fontsize=10)
    ax1r.set_ylim(0, 115)

    ax1.set_title(f"【{location_name}】気温・露点温度（左軸℃）・相対湿度（右軸%） と 各地点の現象コードの関係", fontsize=13)
    ax1.legend(handles=[l1, l2, l3], loc="upper left", fontsize=9)

    _draw_lane_panel(ax2, phenom_df, phenom_cols, location_mapping)
    _finalize_figure(fig, ax1, ax2, times, fig_width)

    fig.savefig(out_path, dpi=SAVE_DPI, bbox_inches="tight")
    plt.close(fig)


def plot_wind_precip(main_df, phenom_df, phenom_cols, location_mapping, location_name, out_path):
    times = main_df["datetime"]
    fig, ax1, ax2, fig_width = _make_fig(times, phenom_cols)

    l1, = ax1.plot(times, main_df["風速(m/s)"], color="#27ae60", linewidth=1.1, label="風速(m/s)", zorder=3)
    ax1.set_ylabel("風速（m/s）", fontsize=10)
    ax1.grid(True, alpha=0.25)

    ax1r = ax1.twinx()
    ax1r.bar(times, main_df["降水量(mm)"], width=0.03, color="#2980b9", alpha=0.55, label="降水量(mm)", zorder=2)
    precip = main_df["降水量(mm)"].dropna()
    pmax = precip.max() if not precip.empty else 1.0
    if not np.isfinite(pmax):
        pmax = 1.0
    ax1r.set_ylim(0, max(pmax * 3.5, 2.0))
    ax1r.set_ylabel("降水量（mm）", fontsize=10)

    ax1.set_title(f"【{location_name}】風速（左軸m/s）・降水量（右軸mm） と 各地点の現象コードの関係", fontsize=13)
    ax1.legend(handles=[l1, Patch(color="#2980b9", alpha=0.55, label="降水量(mm)")], loc="upper left", fontsize=9)

    _draw_lane_panel(ax2, phenom_df, phenom_cols, location_mapping)
    _finalize_figure(fig, ax1, ax2, times, fig_width)

    fig.savefig(out_path, dpi=SAVE_DPI, bbox_inches="tight")
    plt.close(fig)


# ①②のレーンをこれ以上の地点数にしない（読みやすさのため、超える場合は画像を分割する）
MAX_LOCATIONS_PER_MONTHLY_IMAGE = 16


def _cols_with_data_in_month(psub, phenom_cols):
    """その月に実際に記録（「/」も含む）がある地点の列だけを残す。

    記録が無い地点までレーンに描くと、その月の大半が空白のレーンになり読みにくくなる
    （観測期間が月の途中からしかない場合など）。
    """
    return [col for col in phenom_cols if psub[col].map(encode_phenomena_cell).notna().any()]


def _chunk_locations(cols, size):
    return [cols[i:i + size] for i in range(0, len(cols), size)]


def plot_combo_by_month(main_df, phenom_df, phenom_cols, location_mapping, location_name, out_dir):
    months = [(ym, msub, psub) for ym, msub, psub in split_by_month_two(main_df, phenom_df)
             if msub["datetime"].nunique() >= 2]
    generated = []
    progress = _ProgressBar(len(months), prefix="①② 月別グラフ")
    for ym, msub, psub in months:
        progress.update(1, suffix=ym)
        cols_with_data = _cols_with_data_in_month(psub, phenom_cols)
        if not cols_with_data:
            continue
        chunks = _chunk_locations(cols_with_data, MAX_LOCATIONS_PER_MONTHLY_IMAGE)
        for idx, chunk_cols in enumerate(chunks, start=1):
            suffix = f"_{idx}of{len(chunks)}" if len(chunks) > 1 else ""
            loc = (f"{location_name}（{ym}）" if len(chunks) == 1
                  else f"{location_name}（{ym}　{idx}/{len(chunks)}）")
            p1 = os.path.join(out_dir, f"{location_name}_①気温・湿度・露点×現象コード_{ym}{suffix}.png")
            p2 = os.path.join(out_dir, f"{location_name}_②風速・降水量×現象コード_{ym}{suffix}.png")
            plot_temp_humid_dew(msub, psub, chunk_cols, location_mapping, loc, p1)
            plot_wind_precip(msub, psub, chunk_cols, location_mapping, loc, p2)
            generated += [p1, p2]
    progress.close()
    return generated


# ===========================================================================
# 5. 霧予測モデル（地点ごとの多クラス分類）
# ---------------------------------------------------------------------------
# ・「32地点のどこかで霧」ではなく、【地点ごとに個別】のモデルを学習する
# ・「/」(現象なし)+ コード1〜10 の【多クラス分類】として学習する
# ・Open-Meteoから今後の気象予報を取得し、地点ごとに現象コードと霧確率を予測する
# ===========================================================================

FORECAST_LAT = 37.3486   # 予報取得地点の緯度（既定＝只見町付近。--lat / --lon で変更可）
FORECAST_LON = 139.3122  # 予報取得地点の経度
FORECAST_DAYS = 16       # Open-Meteo無料予報で取得できる最大日数
OPEN_METEO_URL = "https://api.open-meteo.com/v1/forecast"

BASE_FEATURES = [
    "気温(℃)", "降水量(mm)", "風速(m/s)", "相対湿度(％)", "露点温度(℃)",
    "気温露点差", "時_sin", "時_cos", "月",
]
# 風向の列がある場合だけ追加される特徴量（select_feature_names()が判定する）
WIND_DIR_FEATURES = ["風向_sin", "風向_cos"]

SUMMARY_FOG_CODES = {1, 2, 3, 4, 5, 6}  # サマリー集計・霧F1評価・霧確率で「霧」とみなすコード

# 地点ごとにモデルを学習するための最低条件（これを満たさない地点はスキップする）
MIN_ROWS_PER_LOCATION = 200
MIN_CLASSES_PER_LOCATION = 2  # 最低2種類以上の現象コードが記録されている必要がある

# 個別モデルの条件を満たさない地点を「全地点共通モデル」で補う（プーリング）ときの条件。
# 個別モデルの下限（MIN_ROWS_PER_LOCATION）はそのままに、データが少ない地点は
# 他地点のデータも借りたモデルで予測できるようにする。
MIN_TOTAL_ROWS_FOR_POOLING = 500   # 共通モデル自体を学習するための全地点合計の最低件数
MIN_ROWS_FOR_POOLED_RESCUE = 30    # 地点ごとの評価ができる最低件数（これ未満は共通モデルでも救わない）

# 霧確率がこの値以上なら「霧」と判定する、というしきい値の探索範囲。
# 既定の50%に根拠はないため、地点ごとに検証期間でF1が最大になる値を選ぶ（`_choose_fog_threshold()`）。
FOG_THRESHOLD_GRID = tuple(round(t, 2) for t in np.arange(0.05, 1.0, 0.05))
DEFAULT_FOG_THRESHOLD = 0.5

# RandomizedSearchCVで探索するハイパーパラメータの範囲（RandomForest）
PARAM_DIST = {
    "model__n_estimators": [200, 300, 500],
    "model__max_depth": [6, 8, 10, 12, 14, None],
    "model__min_samples_leaf": [1, 2, 3, 5],
    "model__max_features": ["sqrt", "log2", None],
}

# 同じく XGBoost（勾配ブースティング）側の探索範囲。
# 木を浅めにして学習率を小さめに振るのがブースティングの定石で、
# 少数クラス（霧）が数十件しかない地点でも過学習しにくくする。
XGB_PARAM_DIST = {
    "model__n_estimators": [200, 400],
    "model__max_depth": [3, 4, 6, 8],
    "model__learning_rate": [0.05, 0.1, 0.2],
    "model__subsample": [0.8, 1.0],
    "model__colsample_bytree": [0.8, 1.0],
    "model__min_child_weight": [1, 3, 5],
}

# ハイブリッドの混合比の候補（RandomForest側の重み）。
# 0.0=XGBoost単独 / 1.0=RandomForest単独 も候補に含めているので、
# 「片方が明らかに良い地点はその片方だけを使う」という判断も自動的に入る。
BLEND_WEIGHTS = [0.0, 0.25, 0.5, 0.75, 1.0]

MODEL_KINDS = ("hybrid", "rf", "xgb")
DEFAULT_MODEL_KIND = "hybrid"
MODEL_KIND_LABELS = {
    "hybrid": "RandomForest + XGBoost のハイブリッド（予測確率を地点ごとの比率で加重平均）",
    "rf": "RandomForest 単独",
    "xgb": "XGBoost 単独",
}


def build_location_class_targets(phenom_df, phenom_cols):
    """地点それぞれについて『現象コード（0="/"、1〜10、欠測はNaN）』を作る。
    地点ごとに独立したラベルであり、他地点の値には一切影響されない。
    """
    codes = phenom_df[phenom_cols].apply(lambda col: col.map(encode_phenomena_cell))
    targets = codes.copy()
    targets.insert(0, "datetime", phenom_df["datetime"].values)
    return targets


# 直近の変化量・移動平均を表す特徴量名（`_add_lag_features()` が作る）
LAG_FEATURE_NAMES = ["気温変化_1h", "気温変化_3h", "風速平均_3h", "降水量_24h合計"]
# 日の出までの残り時間を表す特徴量名（`_add_sunrise_feature()` が作る）
SUN_FEATURE_NAME = "日の出まで残り時間"


def _lag_series(df, col, hours):
    """`col` の `hours` 時間前の値。ちょうどその時刻の実測が無ければNaNになる。

    位置（何行前か）ではなく datetime そのもので引き当てるため、欠測で行が飛んでいたり
    実測と予報のつなぎ目をまたいだりしても、誤った時間差を計算しない。
    """
    s = df.drop_duplicates(subset="datetime").set_index("datetime")[col]
    lookup = pd.DatetimeIndex(df["datetime"]) - pd.Timedelta(hours=hours)
    return pd.Series(s.reindex(lookup).to_numpy(), index=df.index)


def _add_lag_features(df):
    """直近の変化量・移動平均を特徴量として追加する。

    川霧は「その瞬間の気温」よりも「直近でどれだけ気温が下がったか」
    「前日までにどれだけ雨が降ったか」に強く左右されるため、瞬間値だけでは
    表現できない時間的な文脈を補う（現在の主要な精度改善策の一つ）。
    """
    df = df.copy()
    if "気温(℃)" in df.columns:
        df["気温変化_1h"] = df["気温(℃)"] - _lag_series(df, "気温(℃)", 1)
        df["気温変化_3h"] = df["気温(℃)"] - _lag_series(df, "気温(℃)", 3)

    dedup = df.drop_duplicates(subset="datetime").sort_values("datetime")
    dt_index = pd.DatetimeIndex(dedup["datetime"])
    target_index = pd.DatetimeIndex(df["datetime"])

    if "風速(m/s)" in df.columns:
        s = pd.Series(pd.to_numeric(dedup["風速(m/s)"], errors="coerce").to_numpy(), index=dt_index)
        rolled = s.rolling("3h", min_periods=1).mean()
        df["風速平均_3h"] = rolled.reindex(target_index).to_numpy()

    if "降水量(mm)" in df.columns:
        precip = pd.to_numeric(dedup["降水量(mm)"], errors="coerce").fillna(0.0)
        s = pd.Series(precip.to_numpy(), index=dt_index)
        rolled = s.rolling("24h", min_periods=1).sum()
        df["降水量_24h合計"] = rolled.reindex(target_index).to_numpy()
    return df


def _sunrise_hour(ts: pd.Timestamp, lat: float, lon: float) -> float:
    """その日の日の出時刻を、その日の0時(JST)からの経過時間（h）で近似する。

    NOAAの簡易式（標高0m・大気差補正なし）。実用上は数分の誤差があるが、
    「明け方らしさ」を特徴量にするには十分な精度。
    """
    n = ts.timetuple().tm_yday
    lat_rad = math.radians(lat)
    decl = math.radians(23.44) * math.sin(math.radians(360.0 / 365.0 * (n - 81)))
    cos_hour_angle = -math.tan(lat_rad) * math.tan(decl)
    cos_hour_angle = max(-1.0, min(1.0, cos_hour_angle))  # 極域などの白夜/極夜を簡易に丸める
    hour_angle_deg = math.degrees(math.acos(cos_hour_angle))
    solar_noon = 12.0 - lon / 15.0 + 9.0  # JST(UTC+9)における太陽正午の近似
    return solar_noon - hour_angle_deg / 15.0


def _add_sunrise_feature(df, lat=None, lon=None):
    """『日の出まで残り何時間か』を特徴量として追加する（日中は0）。

    川霧は放射冷却で気温が下がりきる明け方に発生しやすいが、日の出時刻は季節で
    大きく変わるため、「月」と「時刻」だけでは季節ごとの日の出の早い/遅いを
    表現しきれない。緯度経度は予報取得地点（既定=只見町付近）で近似する
    （只見川流域は数十km四方に収まるため、地点間の日の出時刻の差は実用上小さい）。
    """
    lat = FORECAST_LAT if lat is None else lat
    lon = FORECAST_LON if lon is None else lon
    df = df.copy()
    dt = pd.to_datetime(df["datetime"])
    day_start = dt.dt.normalize()
    hour_of_day = (dt - day_start).dt.total_seconds().to_numpy() / 3600.0

    unique_days = day_start.unique()
    sunrise_by_day = {d: _sunrise_hour(pd.Timestamp(d), lat, lon) for d in unique_days}
    sunrise_hours = day_start.map(sunrise_by_day).to_numpy(dtype="float64")

    remaining = sunrise_hours - hour_of_day
    df[SUN_FEATURE_NAME] = np.where((remaining < 0) | (remaining > 12), 0.0, remaining)
    return df


def _add_time_features(df):
    """気象データ／予報データに、モデルが使う時間特徴量を追加する共通処理。"""
    df = df.copy()
    df["時"] = df["datetime"].dt.hour
    df["月"] = df["datetime"].dt.month
    df["気温露点差"] = df["気温(℃)"] - df["露点温度(℃)"]
    df["時_sin"] = np.sin(2 * np.pi * df["時"] / 24)
    df["時_cos"] = np.cos(2 * np.pi * df["時"] / 24)
    df = _add_lag_features(df)
    df = _add_sunrise_feature(df)
    if WIND_DIR_LABEL in df.columns:
        # 風向は角度なので、そのまま数値にすると「北(0度)」と「北北西(337.5度)」が
        # 遠い値になってしまう。sin/cosに分解して円周上の近さを保つ。
        rad = np.deg2rad(pd.to_numeric(df[WIND_DIR_LABEL], errors="coerce"))
        # 静穏（角度なし）は sin=cos=0 とし、「風向なし」を表す点として扱う
        df["風向_sin"] = np.sin(rad).fillna(0.0)
        df["風向_cos"] = np.cos(rad).fillna(0.0)
    return df


def select_feature_names(df):
    """そのデータフレームで実際に使える特徴量名のリストを返す。

    気温・降水量などの基本項目に加えて、
    - 風向（列があるときだけ、sin/cosの2列）
    - 日照時間・気圧など任意の気象要素（ファイルにあり、有効値が十分あるときだけ）
    - 直近の変化量・日照時間帯などの時間的な文脈特徴量（`_add_time_features()` が作る）
    を、実際にそのデータフレームで使える分だけ加える。
    """
    features = list(BASE_FEATURES)
    if all(f in df.columns for f in WIND_DIR_FEATURES):
        features += WIND_DIR_FEATURES
    for label in OPTIONAL_MEASURE_KEYWORDS:
        if label in df.columns and df[label].notna().any():
            features.append(label)
    features += [f for f in LAG_FEATURE_NAMES if f in df.columns]
    if SUN_FEATURE_NAME in df.columns:
        features.append(SUN_FEATURE_NAME)
    return features


def _fill_optional_measures(df):
    """任意の気象要素（日照時間・気圧など）の欠測を、その列の中央値で埋める。

    降水量と違って「無ければ0」とは言えない量（気圧など）なので中央値で埋める。
    有効値の割合が低い列はそのままNaNにし、`select_feature_names()` に除外させる
    （dropnaで学習データの大半が消えるのを防ぐため）。
    """
    df = df.copy()
    for label in OPTIONAL_MEASURE_KEYWORDS:
        if label not in df.columns:
            continue
        col = pd.to_numeric(df[label], errors="coerce")
        coverage = col.notna().mean() if len(col) else 0.0
        df[label] = col.fillna(col.median()) if coverage >= OPTIONAL_MEASURE_MIN_COVERAGE else np.nan
    return df


def prepare_features(df, feature_names=None, fill_precip_zero=True):
    """特徴量を組み立てる。

    降水量は「無降水のとき空欄」という記録が多いため、既定では欠測を0mmとして扱う
    （そうしないと dropna でほとんどの行が学習から消えてしまう）。
    """
    out = _add_time_features(df)
    out = _fill_optional_measures(out)
    if fill_precip_zero and "降水量(mm)" in out.columns:
        out["降水量(mm)"] = pd.to_numeric(out["降水量(mm)"], errors="coerce").fillna(0.0)

    # 直近の変化量は、ちょうどN時間前の実測が無い行（データの先頭・欠測の直後など）で
    # NaNになる。学習データが欠測直後にごっそり失われないよう、妥当な既定値で埋める
    # （気温変化=0＝「直近の変化は不明」、風速平均=その時刻の風速、降水量合計=0mm）。
    for name in ("気温変化_1h", "気温変化_3h"):
        if name in out.columns:
            out[name] = out[name].fillna(0.0)
    if "風速平均_3h" in out.columns and "風速(m/s)" in out.columns:
        out["風速平均_3h"] = out["風速平均_3h"].fillna(out["風速(m/s)"])
    if "降水量_24h合計" in out.columns:
        out["降水量_24h合計"] = out["降水量_24h合計"].fillna(0.0)

    if feature_names is None:
        feature_names = select_feature_names(out)
    for name in feature_names:
        if name not in out.columns:
            out[name] = np.nan
    return out, feature_names


def _temporal_train_test_split(X, y, test_ratio=0.2):
    """時系列順を保ったまま、後ろ test_ratio をテストデータにする。"""
    split_idx = max(1, int(len(X) * (1 - test_ratio)))
    if split_idx >= len(X):
        split_idx = len(X) - 1
    return X.iloc[:split_idx], X.iloc[split_idx:], y.iloc[:split_idx], y.iloc[split_idx:]


def _fog_class_f1(y_true, y_pred):
    """霧コード(1〜6)に限定した macro F1。テストに霧が1件も無い場合は NaN。"""
    fog_labels = [c for c in sorted(SUMMARY_FOG_CODES) if c in set(pd.Series(y_true).unique())]
    if not fog_labels:
        return np.nan
    return f1_score(y_true, y_pred, labels=fog_labels, average="macro", zero_division=0)


def _blend_score(y_true, y_pred):
    """モデル選び（探索・混合比）で使う評価値。霧F1を優先し、対象に霧が無ければmacro F1。

    以前は探索の目的関数が f1_macro（全クラス平等）だったため、「/」以外の
    非霧クラス（層雲・雨など）の当たりも霧と同じ重みで最適化されていた。
    霧予測が主目的なので、霧F1を直接の目的関数にする。
    """
    score = _fog_class_f1(y_true, y_pred)
    if np.isnan(score):
        score = f1_score(y_true, y_pred, average="macro", zero_division=0)
    return score


# RandomizedSearchCVの探索目的関数として使う（"f1_macro"のような文字列指定の代わり）。
FOG_F1_SCORER = make_scorer(_blend_score)


def compute_fog_probability(pipe, X):
    """学習済みモデルから「霧（コード1〜6）である確率」を取り出す。

    多クラス分類のpredict()は多数派の「/」を返しがちで、予測グラフが真っ白に
    なりやすい。確率を併せて出すことで「霧になりかけている時間帯」も読み取れる。
    """
    proba = pipe.predict_proba(X)
    classes = list(pipe.classes_)
    fog_idx = [i for i, c in enumerate(classes) if int(c) in SUMMARY_FOG_CODES]
    if not fog_idx:
        return np.zeros(len(X), dtype="float64")
    return proba[:, fog_idx].sum(axis=1)


class _XGBLabelSafeClassifier(ClassifierMixin, BaseEstimator):
    """XGBClassifier を「歯抜けの現象コード」でもそのまま学習できるようにする薄いラッパ。

    XGBoost はラベルが 0,1,2,… と連続していることを要求するが、現象コードは
    地点によって {0,1,2,3,4,7,8,9,10} のように歯抜けになる。fit の中で
    連番へ符号化し、classes_ と predict() は元の現象コードで返すため、
    Pipeline / RandomizedSearchCV / f1_macro をRandomForestと同じ書き方で使える。

    不均衡対策は RandomForest の class_weight="balanced_subsample" に相当する
    sample_weight（balanced）を渡すことで行う。
    """

    def __init__(self, n_estimators=300, max_depth=4, learning_rate=0.1,
                 subsample=1.0, colsample_bytree=1.0, min_child_weight=1,
                 random_state=42):
        self.n_estimators = n_estimators
        self.max_depth = max_depth
        self.learning_rate = learning_rate
        self.subsample = subsample
        self.colsample_bytree = colsample_bytree
        self.min_child_weight = min_child_weight
        self.random_state = random_state

    def fit(self, X, y, sample_weight=None):
        if not _HAS_XGBOOST:
            raise RuntimeError("xgboost が導入されていません")
        y = np.asarray(y)
        self.classes_, encoded = np.unique(y, return_inverse=True)
        self.n_features_in_ = np.asarray(X).shape[1]
        if sample_weight is None:
            sample_weight = compute_sample_weight("balanced", y)
        # 内側のn_jobsは1に固定し、並列化は探索側のn_jobs=-1に任せる
        # （RandomForest側と同じ方針。入れ子並列はスレッド競合で遅くなる）
        self.model_ = XGBClassifier(
            n_estimators=self.n_estimators, max_depth=self.max_depth,
            learning_rate=self.learning_rate, subsample=self.subsample,
            colsample_bytree=self.colsample_bytree,
            min_child_weight=self.min_child_weight,
            random_state=self.random_state, tree_method="hist", n_jobs=1,
            verbosity=0,
        )
        self.model_.fit(X, encoded, sample_weight=sample_weight)
        return self

    def predict_proba(self, X):
        return np.asarray(self.model_.predict_proba(X), dtype="float64")

    def predict(self, X):
        return self.classes_[np.argmax(self.predict_proba(X), axis=1)]


class HybridFogClassifier:
    """RandomForest と XGBoost の予測確率を重み付き平均で混ぜるハイブリッドモデル。

        確率 = weight × RandomForest + (1 - weight) × XGBoost

    weight は地点ごとに検証期間で自動決定する（`_choose_blend_weight()`）。
    weight=1.0 は RandomForest 単独、0.0 は XGBoost 単独と同じ意味になる。

    予測に使う下流のコード（`compute_fog_probability()` /
    `build_location_forecast_codes()`）は classes_ / predict_proba / predict の
    3つしか使わないため、このクラスもその3つだけを提供する。
    """

    def __init__(self, rf=None, xgb=None, weight=1.0):
        if rf is None and xgb is None:
            raise ValueError("rf と xgb の少なくとも一方が必要です")
        if rf is None:
            weight = 0.0
        elif xgb is None:
            weight = 1.0
        self.rf = rf
        self.xgb = xgb
        self.weight = float(weight)
        self.classes_ = np.asarray((rf if rf is not None else xgb).classes_)

    @property
    def model_type(self):
        if self.rf is None:
            return "xgb"
        if self.xgb is None:
            return "rf"
        return "hybrid"

    def describe(self):
        """学習結果の表に出す「採用したモデル」の表示文字列。"""
        if self.model_type == "rf":
            return "RF単独"
        if self.model_type == "xgb":
            return "XGB単独"
        return f"RF:XGB={self.weight:.2f}:{1 - self.weight:.2f}"

    def _aligned_proba(self, model, X):
        """モデルの確率を self.classes_ の並びに合わせて取り出す。"""
        proba = np.asarray(model.predict_proba(X), dtype="float64")
        model_classes = list(np.asarray(model.classes_))
        if model_classes == list(self.classes_):
            return proba
        index = {c: i for i, c in enumerate(self.classes_)}
        out = np.zeros((proba.shape[0], len(self.classes_)), dtype="float64")
        for j, c in enumerate(model_classes):
            if c in index:
                out[:, index[c]] = proba[:, j]
        return out

    def predict_proba(self, X):
        if self.xgb is None:
            return self._aligned_proba(self.rf, X)
        if self.rf is None:
            return self._aligned_proba(self.xgb, X)
        return (self.weight * self._aligned_proba(self.rf, X)
                + (1.0 - self.weight) * self._aligned_proba(self.xgb, X))

    def predict(self, X):
        return self.classes_[np.argmax(self.predict_proba(X), axis=1)]


# ハイパーパラメータ探索に失敗した／データが少なすぎて探索できないときの既定値
RF_FALLBACK_PARAMS = dict(model__n_estimators=500, model__max_depth=14, model__min_samples_leaf=2)
XGB_FALLBACK_PARAMS = dict(model__n_estimators=300, model__max_depth=4, model__learning_rate=0.1)


def _make_base_pipe(kind):
    """RandomForest / XGBoost の素のPipelineを作る。

    どちらも決定木ベースでスケール不変のため StandardScaler は不要。
    PARAM_DIST/set_params の "model__" プレフィックスを使うためPipelineのまま残す。
    """
    if kind == "rf":
        return Pipeline([
            ("model", RandomForestClassifier(class_weight="balanced_subsample", random_state=42)),
        ])
    return Pipeline([("model", _XGBLabelSafeClassifier())])


def _fit_component(kind, X_train, y_train, loc_name):
    """RandomForest または XGBoost を1つ学習する。

    戻り値: (学習済みPipeline, 採用パラメータ, 探索できたか)
    時系列データなのでTimeSeriesSplitを使い、未来データが学習に混ざらないようにする。
    探索の目的関数は霧F1（`FOG_F1_SCORER`）にしている。「/」が大多数の不均衡データでは
    accuracyだと「常に/」でも高得点になり、f1_macroだと霧以外のクラスも霧と同じ重みで
    最適化されてしまうため、霧の当たり具合を直接最適化する。
    """
    label = "RandomForest" if kind == "rf" else "XGBoost"
    param_dist = PARAM_DIST if kind == "rf" else XGB_PARAM_DIST
    fallback_params = RF_FALLBACK_PARAMS if kind == "rf" else XGB_FALLBACK_PARAMS
    n_iter = 10 if kind == "rf" else 8

    cv_splits = min(3, max(2, len(X_train) // 100))
    if cv_splits >= 2 and len(X_train) >= 30:
        try:
            search = RandomizedSearchCV(
                _make_base_pipe(kind), param_dist, n_iter=n_iter,
                cv=TimeSeriesSplit(n_splits=cv_splits),
                scoring=FOG_F1_SCORER, random_state=42, n_jobs=-1, error_score="raise",
            )
            search.fit(X_train, y_train)
            return search.best_estimator_, search.best_params_, True
        except Exception as e:
            print(f"  （{loc_name}: {label}のハイパーパラメータ探索に失敗したため既定値を使います: "
                  f"{type(e).__name__}: {e}）")
    pipe = _make_base_pipe(kind).set_params(**fallback_params)
    pipe.fit(X_train, y_train)
    return pipe, dict(fallback_params), False


def _refit_component(kind, params, X_train, y_train):
    """探索で決まったパラメータのまま、学習データ全体で学習し直す。"""
    pipe = _make_base_pipe(kind).set_params(**params)
    pipe.fit(X_train, y_train)
    return pipe


def _choose_blend_weight(rf, xgb, X_val, y_val):
    """検証期間で最も成績の良い混合比（RandomForest側の重み）を選ぶ。

    戻り値: (重み, {重み: 評価値})
    同点の場合はRandomForest寄り（＝従来どおりの安定側）を採用する。
    """
    scores = {}
    best_w, best_score = 1.0, -1.0
    for w in sorted(BLEND_WEIGHTS, reverse=True):  # RF寄りから順に見る＝同点はRF寄り優先
        score = _blend_score(y_val, HybridFogClassifier(rf, xgb, weight=w).predict(X_val))
        scores[w] = score
        if score > best_score:
            best_w, best_score = w, score
    return best_w, scores


def _choose_fog_threshold(pipe, X_val, y_val):
    """検証期間で、霧確率がこの値以上なら「霧」と判定するしきい値を選ぶ。

    予測現象コード自体（predict()の中身）は変えない。あくまで④のグラフの目安線や
    ⑥CSVの「判定」列で使う、霧確率に対する二値判定のしきい値だけを最適化する。
    検証期間に霧が無い（or 常に霧）ときは、判定のしようがないので既定の50%のままにする。
    """
    if len(X_val) == 0:
        return DEFAULT_FOG_THRESHOLD
    y_bin = pd.Series(y_val).isin(SUMMARY_FOG_CODES).astype(int).to_numpy()
    if y_bin.sum() == 0 or y_bin.sum() == len(y_bin):
        return DEFAULT_FOG_THRESHOLD
    probs = compute_fog_probability(pipe, X_val)
    best_t, best_f1 = DEFAULT_FOG_THRESHOLD, -1.0
    for t in FOG_THRESHOLD_GRID:
        f1 = f1_score(y_bin, (probs >= t).astype(int), zero_division=0)
        if f1 > best_f1:
            best_t, best_f1 = float(t), f1
    return best_t


def _fit_and_select(kind, X_train, y_train, loc_name):
    """指定モデル種別で学習し、検証期間が確保できれば混合比・霧しきい値も選ぶ。

    地点ごとの学習ループと、データが少ない地点を補う「全地点共通モデル」の学習の
    両方から使う共通処理。

    戻り値: (HybridFogClassifier, tuned:bool, threshold:float, rf_pipe, xgb_pipe)
    """
    X_inner, X_val, y_inner, y_val = _temporal_train_test_split(X_train, y_train)
    have_val = y_inner.nunique() >= 2 and len(X_val) >= 10

    rf_pipe = xgb_pipe = None
    if kind == "hybrid":
        if have_val:
            rf_tmp, rf_params, rf_tuned = _fit_component("rf", X_inner, y_inner, loc_name)
            xgb_tmp, xgb_params, xgb_tuned = _fit_component("xgb", X_inner, y_inner, loc_name)
            weight, _ = _choose_blend_weight(rf_tmp, xgb_tmp, X_val, y_val)
            threshold = _choose_fog_threshold(HybridFogClassifier(rf_tmp, xgb_tmp, weight), X_val, y_val)
            # 比率としきい値が決まったら、検証期間も含めた学習データ全体で学習し直す
            # （探索はしないので追加の計算コストは小さい）。
            rf_pipe = _refit_component("rf", rf_params, X_train, y_train)
            xgb_pipe = _refit_component("xgb", xgb_params, X_train, y_train)
            tuned = rf_tuned and xgb_tuned
        else:
            # 検証期間を切り出せないほど小さい地点は、比率としきい値を既定のままにする
            rf_pipe, _, rf_tuned = _fit_component("rf", X_train, y_train, loc_name)
            xgb_pipe, _, xgb_tuned = _fit_component("xgb", X_train, y_train, loc_name)
            weight, tuned, threshold = 0.5, rf_tuned and xgb_tuned, DEFAULT_FOG_THRESHOLD
    else:
        single_kind = kind  # "rf" または "xgb"
        weight = 1.0 if single_kind == "rf" else 0.0
        if have_val:
            tmp, params, s_tuned = _fit_component(single_kind, X_inner, y_inner, loc_name)
            tmp_pipe = (HybridFogClassifier(tmp, None, weight=1.0) if single_kind == "rf"
                       else HybridFogClassifier(None, tmp, weight=0.0))
            threshold = _choose_fog_threshold(tmp_pipe, X_val, y_val)
            final_pipe = _refit_component(single_kind, params, X_train, y_train)
            tuned = s_tuned
        else:
            final_pipe, _, tuned = _fit_component(single_kind, X_train, y_train, loc_name)
            threshold = DEFAULT_FOG_THRESHOLD
        if single_kind == "rf":
            rf_pipe = final_pipe
        else:
            xgb_pipe = final_pipe

    pipe = HybridFogClassifier(rf_pipe, xgb_pipe, weight=weight)
    return pipe, tuned, threshold, rf_pipe, xgb_pipe


def _collect_location_xy(merged, col, dummy_blocks, feature_names):
    """1地点ぶんの学習データ（実測がある行だけ・ダミー期間を除外）を取り出す。"""
    sub = merged.dropna(subset=[col]).sort_values("datetime").reset_index(drop=True)
    for start, end, _, _ in dummy_blocks.get(col, []):
        sub = sub[(sub["datetime"] < start) | (sub["datetime"] > end)]
    sub = sub.reset_index(drop=True)
    return sub, sub[feature_names], sub[col].astype(int)


def _top_feature_importances(rf_pipe, xgb_pipe, weight, feature_names, top_n=3):
    """学習済みモデルで重要度の高かった特徴量を上位から返す（`[(名前, 重要度), ...]`）。

    ハイブリッドの場合はRF/XGBの重要度を混合比で加重平均する
    （どちらも同じ feature_names の並びで学習しているため、そのまま足せる）。
    """
    imps = np.zeros(len(feature_names))
    if rf_pipe is not None:
        imps = imps + weight * rf_pipe.named_steps["model"].feature_importances_
    if xgb_pipe is not None:
        imps = imps + (1 - weight) * xgb_pipe.named_steps["model"].model_.feature_importances_
    order = np.argsort(imps)[::-1][:top_n]
    return [(feature_names[i], float(imps[i])) for i in order]


def _resolve_model_kind(model):
    """指定されたモデル種別を検証し、xgboostが無い環境ではrfへ降格させる。"""
    kind = (model or DEFAULT_MODEL_KIND).lower()
    if kind not in MODEL_KINDS:
        raise ValueError(f"model は {MODEL_KINDS} のいずれかを指定してください（指定値: {model}）")
    if kind in ("hybrid", "xgb") and not _HAS_XGBOOST:
        print("  ※ xgboost が見つからないため RandomForest 単独で学習します"
              "（`pip install xgboost` で有効になります。Google Colabには標準で入っています）")
        return "rf"
    return kind


def train_location_models(main_df, phenom_df, phenom_cols, location_mapping,
                          model=DEFAULT_MODEL_KIND):
    """地点それぞれについて、独立した多クラス分類モデルを学習する。

    model="hybrid"（既定）では RandomForest と XGBoost を両方学習し、
    予測確率を地点ごとの最適な比率で混ぜ合わせる（`HybridFogClassifier`）。
    model="rf" / "xgb" ならその片方だけを使う。

    個別の学習条件（`MIN_ROWS_PER_LOCATION`件・`MIN_CLASSES_PER_LOCATION`種類）を
    満たさない地点は、条件を満たす地点も含めた全地点共通のデータで学習した
    「共通モデル」で補う（`model_type` が "pooled(...)" になる）。

    戻り値: {列文字: {"pipe": HybridFogClassifier, "accuracy": float, "f1": float,
                      "fog_f1": float, "n": int, "n_classes": int, "classes": list,
                      "features": list, "model_type": str, "blend_weight": float,
                      "fog_threshold": float, "rf_fog_f1": float, "xgb_fog_f1": float,
                      "pooled": bool}}
    """
    targets = build_location_class_targets(phenom_df, phenom_cols)
    merged = pd.merge(main_df, targets, on="datetime", how="inner")
    merged, feature_names = prepare_features(merged)
    n_before = len(merged)
    merged = merged.dropna(subset=feature_names)

    print("\n" + "=" * 66)
    print("■ 地点ごとの現象コード予測モデルを学習（多クラス分類）")
    print("=" * 66)
    kind = _resolve_model_kind(model)
    print(f"  学習方法      : {MODEL_KIND_LABELS[kind]}")
    print(f"  使用する特徴量: {', '.join(feature_names)}")
    if n_before - len(merged):
        print(f"  気象要素が欠測のため学習対象から外した行: {n_before - len(merged)}件 / {n_before}件")

    dummy_blocks_by_col = {col: find_dummy_blocks(phenom_df, col) for col in phenom_cols}
    all_warnings = []
    for col, blocks in dummy_blocks_by_col.items():
        loc_name = location_mapping.get(col, col)
        for start, end, ratio, n in blocks:
            all_warnings.append(
                f"{loc_name}: {start.strftime('%Y/%m/%d')}〜{end.strftime('%Y/%m/%d')}（{n}件）は"
                f"「/」の割合が{ratio:.1%}と極端に低く、ダミーデータの可能性があるため学習から除外します。"
            )
    if all_warnings:
        print("\n【データ品質チェック】以下の期間はダミーデータの可能性があるため、学習から除外します:")
        for w in all_warnings:
            print(f"  ⚠ {w}")

    print()
    progress = _ProgressBar(len(phenom_cols), prefix="地点ごとの学習")
    models = {}
    learned_rows = []   # 表示用: (loc_name, n, n_classes, acc, f1, fog_f1, desc, threshold, tag)
    skipped_rows = []   # 表示用: (loc_name, n, reason)
    for col in phenom_cols:
        loc_name = location_mapping.get(col, col)
        progress.update(1, suffix=loc_name)
        sub, X, y = _collect_location_xy(merged, col, dummy_blocks_by_col, feature_names)

        if len(sub) < MIN_ROWS_PER_LOCATION:
            reason = "ダミー判定で全期間を除外" if len(sub) == 0 else "データ不足でスキップ"
            skipped_rows.append((loc_name, len(sub), reason))
            continue

        n_classes = y.nunique()
        if n_classes < MIN_CLASSES_PER_LOCATION:
            skipped_rows.append((loc_name, len(sub), "現象の種類が少なくスキップ"))
            continue

        X_train, X_test, y_train, y_test = _temporal_train_test_split(X, y)
        if y_train.nunique() < 2:
            skipped_rows.append((loc_name, len(sub), "分割後クラス不足でスキップ"))
            continue

        pipe, tuned, threshold, rf_pipe, xgb_pipe = _fit_and_select(kind, X_train, y_train, loc_name)

        y_pred = pipe.predict(X_test)
        acc = accuracy_score(y_test, y_pred)
        f1 = f1_score(y_test, y_pred, average="weighted", zero_division=0)
        fog_f1 = _fog_class_f1(y_test, y_pred)
        # 「ハイブリッドにして本当に良くなったのか」を利用者が確かめられるよう、
        # 混ぜる前のRandomForest単独・XGBoost単独の霧F1も同じテスト期間で測っておく。
        rf_fog_f1 = _fog_class_f1(y_test, rf_pipe.predict(X_test)) if rf_pipe is not None else np.nan
        xgb_fog_f1 = _fog_class_f1(y_test, xgb_pipe.predict(X_test)) if xgb_pipe is not None else np.nan

        # テスト期間に現象が1種類しかない場合、正解率は「常に多数派を答える」だけで
        # 100%になってしまう。数字を鵜呑みにしないよう、その旨をここで明示する。
        eval_trivial = y_test.nunique() < 2
        tag = "調整済" if tuned else "既定値"
        if eval_trivial:
            tag += "・評価期間に現象1種のみ＝正解率は参考値"
        learned_rows.append((loc_name, len(sub), n_classes, acc, f1, fog_f1, pipe.describe(),
                             threshold, tag))
        models[col] = {
            "pipe": pipe, "accuracy": acc, "f1": f1, "fog_f1": fog_f1,
            "n": len(sub), "n_classes": n_classes, "classes": sorted(y.unique().tolist()),
            "features": feature_names, "eval_trivial": eval_trivial,
            "model_type": pipe.model_type, "blend_weight": pipe.weight,
            "fog_threshold": threshold, "rf_fog_f1": rf_fog_f1, "xgb_fog_f1": xgb_fog_f1,
            "pooled": False, "rf_pipe": rf_pipe, "xgb_pipe": xgb_pipe, "loc_name": loc_name,
        }
    progress.close()

    _print_training_table(learned_rows, skipped_rows, len(phenom_cols))

    if models:
        accs = [m["accuracy"] for m in models.values()]
        scored = [m for m in models.values() if not np.isnan(m["fog_f1"])]
        fog_f1s = [m["fog_f1"] for m in scored]
        print(f"\n学習できた地点数: {len(models)} / {len(phenom_cols)}　平均正解率: {np.mean(accs):.1%}", end="")
        if fog_f1s:
            print(f"　平均霧F1: {np.mean(fog_f1s):.3f}")
        else:
            print("　（テストに霧が含まれる地点がなく、霧F1は算出不可）")
        if kind == "hybrid" and scored:
            # ハイブリッド化の効果をその場で確認できるようにする
            rf_only = np.mean([m["rf_fog_f1"] for m in scored])
            xgb_only = np.mean([m["xgb_fog_f1"] for m in scored])
            print(f"　平均霧F1の内訳 → RandomForest単独: {rf_only:.3f} / XGBoost単独: {xgb_only:.3f}"
                  f" / ハイブリッド: {np.mean(fog_f1s):.3f}（{len(scored)}地点で比較）")
        print("※ 正解率は「/」（現象なし）が大多数だと高く出ます。霧の当たり具合は霧F1で確認してください。")
        _print_top_features(models, feature_names)
    else:
        print("\n【警告】どの地点も学習条件を満たさず、モデルを1つも作れませんでした。")

    _rescue_with_pooled_model(models, merged, phenom_cols, location_mapping,
                              dummy_blocks_by_col, feature_names, kind)

    for m in models.values():
        m.pop("rf_pipe", None)
        m.pop("xgb_pipe", None)
    return models


def _print_training_table(learned_rows, skipped_rows, n_total):
    """学習できた地点の表と、スキップした地点の要約を分けて表示する。

    以前は「学習できた地点」と「スキップした地点」を同じ表に混ぜ、地点名の表示幅も
    文字数基準で揃えていたため、全角文字数の違う地点名が並ぶと桁がずれていた。
    `_pad()`（表示幅基準）で揃え、見やすさのために2つの表に分ける。
    """
    name_w = max([12] + [_disp_width(r[0]) for r in learned_rows + skipped_rows])

    if learned_rows:
        print(f"【学習できた地点】 {len(learned_rows)} / {n_total}地点")
        header = (f"  {_pad('地点名', name_w)} {_pad('件数', 7, '>')} {_pad('クラス数', 8, '>')} "
                  f"{_pad('正解率', 7, '>')} {_pad('F1', 6, '>')} {_pad('霧F1', 6, '>')} "
                  f"{_pad('霧判定閾値', 10, '>')}  採用モデル")
        print(header)
        print("  " + "-" * (_disp_width(header) - 2))
        for loc_name, n, n_classes, acc, f1, fog_f1, desc, threshold, tag in learned_rows:
            fog_f1_str = f"{fog_f1:.3f}" if not np.isnan(fog_f1) else "-"
            print(f"  {_pad(loc_name, name_w)} {n:>7} {n_classes:>8} {acc:>7.1%} {f1:>6.3f} "
                  f"{_pad(fog_f1_str, 6, '>')} {_pad(f'{threshold:.0%}', 10, '>')}  {desc}（{tag}）")

    if skipped_rows:
        print(f"\n【スキップした地点】 {len(skipped_rows)} / {n_total}地点"
              "（後述の「共通モデル」で補えるか確認します）")
        by_reason = {}
        for loc_name, n, reason in skipped_rows:
            by_reason.setdefault(reason, []).append((loc_name, n))
        for reason, items in by_reason.items():
            names = "、".join(f"{name}({n}件)" for name, n in items)
            print(f"  {reason}（{len(items)}地点）: {names}")


def _print_top_features(models, feature_names):
    """地点ごとに効いている特徴量の上位3つを表示する（予測を信じてよいかの参考情報）。"""
    real_models = {col: m for col, m in models.items() if not m.get("pooled")}
    if not real_models:
        return
    print("\n【参考】地点ごとに効いている特徴量トップ3（重要度）:")
    name_w = max(_disp_width(m.get("loc_name", col)) for col, m in real_models.items())
    for col, m in real_models.items():
        top = _top_feature_importances(m.get("rf_pipe"), m.get("xgb_pipe"), m["blend_weight"],
                                       feature_names)
        text = "  ".join(f"{name}({imp:.2f})" for name, imp in top)
        loc_name = m.get("loc_name", col)
        print(f"  {_pad(loc_name, name_w)}  {text}")


def _rescue_with_pooled_model(models, merged, phenom_cols, location_mapping,
                              dummy_blocks_by_col, feature_names, kind):
    """個別の学習条件を満たさなかった地点を、全地点共通のデータで学習したモデルで補う。

    只見川流域の32地点は同じ気象条件の影響を受けるため、データが少ない地点でも
    「気象条件と現象コードの関係」自体は他地点からある程度借りられる、という考え方。
    共通モデルは地点を区別しない（同じ気象条件なら同じ予測になる）ため、
    地点固有の霧の出やすさまでは学習できない点には注意（README参照）。
    """
    skipped_cols = [c for c in phenom_cols if c not in models]
    if not skipped_cols:
        return

    pool_frames = []
    for col in phenom_cols:
        sub, X, y = _collect_location_xy(merged, col, dummy_blocks_by_col, feature_names)
        if sub.empty:
            continue
        part = X.copy()
        part["__target"] = y.to_numpy()
        part["__datetime"] = sub["datetime"].to_numpy()
        pool_frames.append(part)

    if not pool_frames:
        return
    pooled = pd.concat(pool_frames, ignore_index=True).sort_values("__datetime").reset_index(drop=True)
    if len(pooled) < MIN_TOTAL_ROWS_FOR_POOLING or pooled["__target"].nunique() < 2:
        print(f"\n【共通モデル】データ不足の{len(skipped_cols)}地点がありますが、"
              f"全地点を合わせても学習データが少なく（{len(pooled)}件）、共通モデルも作れませんでした。")
        return

    Xp, yp = pooled[feature_names], pooled["__target"]
    Xp_train, Xp_test, yp_train, yp_test = _temporal_train_test_split(Xp, yp)
    if yp_train.nunique() < 2:
        return

    print("\n" + "-" * 66)
    print(f"■ データ不足の{len(skipped_cols)}地点を「全地点共通モデル」で補います"
          f"（学習に使う合計件数: {len(pooled)}件、{MODEL_KIND_LABELS[kind]}）")
    pooled_pipe, pooled_tuned, pooled_threshold, *_ = _fit_and_select(kind, Xp_train, yp_train, "共通モデル")

    rescued = []
    for col in skipped_cols:
        sub, X, y = _collect_location_xy(merged, col, dummy_blocks_by_col, feature_names)
        if len(sub) < MIN_ROWS_FOR_POOLED_RESCUE or y.nunique() < 2:
            continue
        X_train2, X_test2, y_train2, y_test2 = _temporal_train_test_split(X, y)
        if len(X_test2) == 0:
            continue
        y_pred2 = pooled_pipe.predict(X_test2)
        acc2 = accuracy_score(y_test2, y_pred2)
        f12 = f1_score(y_test2, y_pred2, average="weighted", zero_division=0)
        fog_f12 = _fog_class_f1(y_test2, y_pred2)
        loc_name = location_mapping.get(col, col)
        models[col] = {
            "pipe": pooled_pipe, "accuracy": acc2, "f1": f12, "fog_f1": fog_f12,
            "n": len(sub), "n_classes": int(y.nunique()), "classes": sorted(y.unique().tolist()),
            "features": feature_names, "eval_trivial": y_test2.nunique() < 2,
            "model_type": f"pooled({pooled_pipe.model_type})", "blend_weight": pooled_pipe.weight,
            "fog_threshold": pooled_threshold, "rf_fog_f1": np.nan, "xgb_fog_f1": np.nan,
            "pooled": True, "loc_name": loc_name,
        }
        rescued.append((loc_name, len(sub), fog_f12))

    if rescued:
        name_w = max(_disp_width(n) for n, _, _ in rescued)
        print(f"  共通モデルで学習できた地点: {len(rescued)}地点"
              f"（{'調整済' if pooled_tuned else '既定値'}、霧判定閾値={pooled_threshold:.0%}）")
        for loc_name, n, fog_f12 in rescued:
            fog_f1_str = f"{fog_f12:.3f}" if not np.isnan(fog_f12) else "-"
            print(f"    {_pad(loc_name, name_w)} {n:>6}件  霧F1={fog_f1_str}")
        print("  ※ 共通モデルは地点を区別しないため、地点固有の霧の出やすさは反映されません。"
              "参考程度に確認してください。")
    else:
        print("  共通モデルで評価できるほどのデータがある地点は、他にありませんでした。")


# Open-Meteoの変数名 → 学習データと同じ列名・単位への変換。
# 任意項目（日照時間・気圧・日射量・視程）も常にまとめて取得しておき、学習データ側に
# その項目が実際にあった地点だけ特徴量として使う（`select_feature_names()` が判定）。
#   sunshine_duration   : 秒 → 時間（÷3600）
#   shortwave_radiation : W/m²（1時間平均）→ MJ/m²（×0.0036）
#   visibility          : m → km（÷1000）
#   surface_pressure    : hPa（変換不要）
_FORECAST_VARS = {
    "temperature_2m": ("気温(℃)", None),
    "precipitation": ("降水量(mm)", None),
    "relative_humidity_2m": ("相対湿度(％)", None),
    "dew_point_2m": ("露点温度(℃)", None),
    "wind_speed_10m": ("風速(m/s)", None),
    "wind_direction_10m": (WIND_DIR_LABEL, None),
    "sunshine_duration": ("日照時間(時間)", lambda v: v / 3600.0),
    "surface_pressure": ("気圧(hPa)", None),
    "shortwave_radiation": ("日射量(MJ/m2)", lambda v: v * 0.0036),
    "visibility": ("視程(km)", lambda v: v / 1000.0),
}


def _forecast_payload_to_df(hourly: dict) -> pd.DataFrame:
    """Open-Meteoの `hourly` ブロック1つを、学習データと同じ列名のDataFrameに変換する。"""
    data = {"datetime": pd.to_datetime(hourly["time"])}
    for var, (label, convert) in _FORECAST_VARS.items():
        if var not in hourly:
            continue
        values = pd.to_numeric(pd.Series(hourly[var]), errors="coerce")
        data[label] = convert(values) if convert else values
    df = pd.DataFrame(data)
    if "降水量(mm)" in df.columns:
        df["降水量(mm)"] = df["降水量(mm)"].fillna(0.0)

    # 予報値に欠測が混じっているとモデルが例外を出すため、時間方向に補間して埋める
    numeric_cols = [c for c in df.columns if c != "datetime"]
    df[numeric_cols] = df[numeric_cols].interpolate(limit_direction="both")
    df = df.dropna(subset=["気温(℃)", "相対湿度(％)", "露点温度(℃)", "風速(m/s)"]).reset_index(drop=True)
    if df.empty:
        raise ValueError("Open-Meteoから有効な予報値を取得できませんでした。")
    return df


def fetch_forecast_range(lat: float = None, lon: float = None, forecast_days: int = None):
    """Open-Meteo（無料・APIキー不要）から今日以降の時間別予報を1回のリクエストで取得する。

    風速は既定だとkm/hで返ってくる。学習データ（気象庁）はm/sなので、
    wind_speed_unit=ms を必ず指定して単位を揃える（ここが揃っていないと
    予報の風速が実際の約3.6倍になり、予測が大きく狂う）。
    """
    lat = FORECAST_LAT if lat is None else lat
    lon = FORECAST_LON if lon is None else lon
    forecast_days = FORECAST_DAYS if forecast_days is None else forecast_days

    params = {
        "latitude": lat,
        "longitude": lon,
        "hourly": ",".join(_FORECAST_VARS.keys()),
        "timezone": "Asia/Tokyo",
        "forecast_days": forecast_days,
        "wind_speed_unit": "ms",      # 気象庁データ（m/s）に合わせる
        "precipitation_unit": "mm",
        "temperature_unit": "celsius",
    }
    resp = requests.get(OPEN_METEO_URL, params=params, timeout=20)
    resp.raise_for_status()
    payload = resp.json()
    if "hourly" not in payload:
        raise ValueError(f"Open-Meteoの応答に hourly が含まれていません: {payload}")
    return _forecast_payload_to_df(payload["hourly"])


def load_location_coordinates(filepath):
    """『地点名,緯度,経度』形式のCSVを読み、{地点名: (緯度, 経度)} を返す。

    地点ごとに予報を取得したいときに使う（既定は全地点で1地点分の予報を共用している）。
    列名は「地点」「緯度」「経度」を含んでいれば順不同で構わない。
    """
    df = pd.read_csv(filepath)
    df.columns = [str(c).strip() for c in df.columns]
    name_col = next((c for c in df.columns if "地点" in c or "name" in c.lower()), df.columns[0])
    lat_col = next((c for c in df.columns if "緯度" in c or "lat" in c.lower()), None)
    lon_col = next((c for c in df.columns if "経度" in c or "lon" in c.lower()), None)
    if lat_col is None or lon_col is None:
        raise ValueError(
            "緯度・経度の列が見つかりません。列名に「緯度」「経度」を含む列を用意してください"
            "（例: 地点名,緯度,経度）。"
        )
    coords = {}
    for _, row in df.iterrows():
        name = str(row[name_col]).strip()
        try:
            coords[name] = (float(row[lat_col]), float(row[lon_col]))
        except (TypeError, ValueError):
            continue
    return coords


def fetch_forecast_multi(coords: dict, forecast_days: int = None):
    """複数地点の予報を1回のリクエストでまとめて取得する（Open-Meteoの緯度経度カンマ区切り機能）。

    戻り値: {地点名: 予報DataFrame}（`fetch_forecast_range()` と同じ列構成）
    """
    if not coords:
        return {}
    forecast_days = FORECAST_DAYS if forecast_days is None else forecast_days
    names = list(coords.keys())
    params = {
        "latitude": ",".join(str(coords[n][0]) for n in names),
        "longitude": ",".join(str(coords[n][1]) for n in names),
        "hourly": ",".join(_FORECAST_VARS.keys()),
        "timezone": "Asia/Tokyo",
        "forecast_days": forecast_days,
        "wind_speed_unit": "ms",
        "precipitation_unit": "mm",
        "temperature_unit": "celsius",
    }
    resp = requests.get(OPEN_METEO_URL, params=params, timeout=30)
    resp.raise_for_status()
    payload = resp.json()
    # 1地点だけのときはdict、複数地点のときはlistで返ってくる（Open-Meteoの仕様）
    items = [payload] if isinstance(payload, dict) else payload
    result = {}
    for name, item in zip(names, items):
        if "hourly" not in item:
            continue
        result[name] = _forecast_payload_to_df(item["hourly"])
    return result


def build_location_forecast_codes(models, forecast_raw, forecast_by_col: dict = None):
    """予報データに対して、地点ごとに現象コードと霧確率を予測する。

    forecast_by_col を渡すと、その列だけは対応する専用の予報（地点ごとの緯度経度で
    取得したもの）を使う。渡されない／該当が無い地点は forecast_raw（既定の1地点分の予報）
    を使う。

    戻り値: (pred_df, prob_df)
      pred_df : datetime列 + 各地点の予測コード（0〜10）
      prob_df : datetime列 + 各地点の霧（コード1〜6）確率（0〜1）
    """
    forecast_by_col = forecast_by_col or {}
    pred_df = pd.DataFrame({"datetime": forecast_raw["datetime"].values})
    prob_df = pd.DataFrame({"datetime": forecast_raw["datetime"].values})
    for col, info in models.items():
        raw = forecast_by_col.get(col, forecast_raw)
        feature_names = info.get("features") or select_feature_names(raw)
        df, feature_names = prepare_features(raw, feature_names)
        X = df[feature_names]
        pred_df[col] = info["pipe"].predict(X).astype(int)
        prob_df[col] = compute_fog_probability(info["pipe"], X)
    return pred_df, prob_df


def get_last_observed_datetime(phenom_df, col):
    """その地点について、実際に結果が入力されている最後の日時を返す。
    未来日付の空欄行がファイルにあっても無視する。
    """
    codes = phenom_df[col].map(encode_phenomena_cell)
    valid_dates = phenom_df.loc[codes.notna(), "datetime"]
    if valid_dates.empty:
        return phenom_df["datetime"].max()
    return valid_dates.max()


def find_dummy_blocks(phenom_df, col, min_block_rows: int = 50, gap_hours: float = 24 * 7,
                      max_slash_ratio: float = 0.02):
    """観測データを時間的な『まとまり(ブロック)』に分割し、その中で「/」(現象なし)が
    極端に少ない（＝ダミーデータの疑いがある）ブロックを検出する。

    戻り値: [(開始日時, 終了日時, 「/」の割合, 件数), ...]
    """
    codes = phenom_df[col].map(encode_phenomena_cell)
    valid_mask = codes.notna()
    valid_df = pd.DataFrame({
        "datetime": phenom_df.loc[valid_mask, "datetime"].values,
        "code": codes[valid_mask].values,
    })
    if valid_df.empty:
        return []

    # 時間差がgap_hours以上空いたら別ブロックとみなす
    time_diff = valid_df["datetime"].diff().dt.total_seconds() / 3600
    block_id = (time_diff > gap_hours).cumsum()

    suspicious = []
    for _, block in valid_df.groupby(block_id):
        if len(block) < min_block_rows:
            continue
        slash_ratio = (block["code"] == 0).mean()
        if slash_ratio <= max_slash_ratio:
            suspicious.append((block["datetime"].min(), block["datetime"].max(), slash_ratio, len(block)))
    return suspicious


# ===========================================================================
# 6. 予測結果のグラフ・CSV出力
# ===========================================================================

FORECAST_MIN_FIG_WIDTH = 16
FORECAST_MAX_FIG_WIDTH = 34
FORECAST_INCH_PER_DAY = 1.2
# 実測の最終記録と予報開始がこれ以上離れていたら、横軸を分割して描く
AXIS_BREAK_GAP_DAYS = 2.0


def _apply_date_axis(ax, span_days, label_all=True):
    day_interval = 1 if span_days <= 30 else max(1, int(span_days // 30))
    ax.xaxis.set_major_locator(mdates.DayLocator(interval=day_interval))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y/%m/%d"))
    ax.grid(True, axis="x", which="major", color="gray", alpha=0.25, linewidth=0.6, zorder=0)
    ax.tick_params(axis="x", labelbottom=label_all, labelsize=8, labelrotation=45)


def _draw_axis_break_marks(ax_left, ax_right):
    """分割した横軸のつなぎ目に「//」の切断記号を描く。"""
    kw = dict(marker=[(-1, -0.6), (1, 0.6)], markersize=8, linestyle="none",
              color="dimgray", mec="dimgray", mew=1.2, clip_on=False)
    ax_left.plot([1, 1], [0, 1], transform=ax_left.transAxes, **kw)
    ax_right.plot([0, 0], [0, 1], transform=ax_right.transAxes, **kw)


def plot_single_location_forecast(main_df, phenom_df, col, location_mapping,
                                  forecast_raw, pred_series, location_name, out_dir,
                                  history_days: int = 5, prob_series=None,
                                  fog_threshold: float = DEFAULT_FOG_THRESHOLD):
    """1地点だけの「実測（直近history_days日）＋ 今後の予測」グラフを作る。

    上段  : 気温・露点・湿度（実測＝実線／予報＝点線、色は実測と同じ）
    中段  : 霧（コード1〜6）の予測確率
    下段  : 現象コード（実測＋予測、①②と同じ配色）

    実測の終点は「ファイルの最後の行」ではなく【実際に結果が入力されている最後の時刻】を使う。
    実測の終わりと予報の開始が離れている場合（記録の入力が遅れている場合など）は、
    間の空白をそのまま描くと図が極端に横長になって読めなくなるため、
    横軸を「実測期間」と「予報期間」に分割して描く（つなぎ目に // を表示）。
    """
    loc_name = location_mapping.get(col, col)

    last_observed = get_last_observed_datetime(phenom_df, col)
    hist_cutoff = last_observed - pd.Timedelta(days=history_days)
    hist_main = main_df[(main_df["datetime"] >= hist_cutoff) &
                        (main_df["datetime"] <= last_observed)].reset_index(drop=True)
    hist_phenom = phenom_df[(phenom_df["datetime"] >= hist_cutoff) &
                            (phenom_df["datetime"] <= last_observed)].reset_index(drop=True)
    hist_codes = hist_phenom[col].map(encode_phenomena_cell).to_numpy(dtype="float64")

    fc_start = forecast_raw["datetime"].min()
    fc_end = forecast_raw["datetime"].max()
    hist_start = hist_main["datetime"].min() if len(hist_main) else hist_cutoff
    hist_end = hist_main["datetime"].max() if len(hist_main) else last_observed

    gap_days = (fc_start - hist_end).total_seconds() / 86400.0
    if len(hist_main) and gap_days > AXIS_BREAK_GAP_DAYS:
        segments = [(hist_start, hist_end), (fc_start, fc_end)]
    elif len(hist_main):
        segments = [(min(hist_start, fc_start), max(hist_end, fc_end))]
    else:
        segments = [(fc_start, fc_end)]

    seg_days = [max(0.5, (e - s).total_seconds() / 86400.0) for s, e in segments]
    fig_width = float(np.clip(sum(seg_days) * FORECAST_INCH_PER_DAY,
                              FORECAST_MIN_FIG_WIDTH, FORECAST_MAX_FIG_WIDTH))

    fig, axes = plt.subplots(
        3, len(segments), figsize=(fig_width, 8.2), squeeze=False,
        sharey="row",
        gridspec_kw={"height_ratios": [5, 1.6, 1.2], "width_ratios": seg_days,
                     "hspace": 0.12, "wspace": 0.05},
    )

    probs = None if prob_series is None else np.asarray(prob_series, dtype="float64") * 100.0
    pred_codes = np.asarray(pred_series, dtype="int64")
    x_hist = mdates.date2num(hist_phenom["datetime"].to_numpy()) if len(hist_phenom) else np.array([])
    x_fore = mdates.date2num(forecast_raw["datetime"].to_numpy())
    n_days = max(1, round(len(forecast_raw) / 24))

    humid_axes = []
    for j, (seg_start, seg_end) in enumerate(segments):
        ax1, axp, ax2 = axes[0][j], axes[1][j], axes[2][j]
        is_forecast_panel = seg_end >= fc_start
        is_last = (j == len(segments) - 1)

        # ーーー 上段：気温・露点温度（左軸）・相対湿度（右軸） ーーー
        # 各項目は実測・予報とも同じ色を使い、実線=実測／点線=予報で区別する。
        # 「ここから先が予測」という点は背景の網掛けと境界線で示す。
        if is_forecast_panel:
            for ax in (ax1, axp, ax2):
                ax.axvspan(fc_start, fc_end, color="gray", alpha=0.08, zorder=0)
                ax.axvline(fc_start, color="dimgray", linewidth=1.5, linestyle=":")

        ax1.plot(hist_main["datetime"], hist_main["気温(℃)"], color="#e74c3c", linewidth=1.2,
                 label="気温(℃) [実測]")
        ax1.plot(hist_main["datetime"], hist_main["露点温度(℃)"], color="#16a085", linewidth=1.2,
                 label="露点温度(℃) [実測]")
        ax1.plot(forecast_raw["datetime"], forecast_raw["気温(℃)"], color="#e74c3c", linewidth=1.5,
                 linestyle="--", label="気温(℃) [予報]")
        ax1.plot(forecast_raw["datetime"], forecast_raw["露点温度(℃)"], color="#16a085", linewidth=1.5,
                 linestyle="--", label="露点温度(℃) [予報]")
        ax1.grid(True, axis="y", alpha=0.25)

        ax1r = ax1.twinx()
        humid_axes.append(ax1r)
        ax1r.plot(hist_main["datetime"], hist_main["相対湿度(％)"], color="#8e44ad", linewidth=0.9,
                  alpha=0.6, label="相対湿度(％) [実測]")
        ax1r.plot(forecast_raw["datetime"], forecast_raw["相対湿度(％)"], color="#8e44ad", linewidth=1.2,
                  alpha=0.9, linestyle="--", label="相対湿度(％) [予報]")
        ax1r.set_ylim(0, 115)

        # ーーー 中段：霧（コード1〜6）の予測確率 ーーー
        # 目安線は固定50%ではなく、地点ごとに検証期間で最適化したしきい値を使う
        # （`fog_threshold`。地点によって霧の出やすさ・データの偏りが違うため、
        # 一律50%だと見逃しや誤警戒が多くなる地点があった）。
        threshold_pct = fog_threshold * 100.0
        if probs is not None:
            axp.fill_between(forecast_raw["datetime"], 0, probs, color="#2980b9", alpha=0.30, zorder=2)
            axp.plot(forecast_raw["datetime"], probs, color="#1a5276", linewidth=1.4, zorder=3,
                     label="霧(1〜6)の予測確率")
            axp.axhline(threshold_pct, color="#c0392b", linewidth=0.9, linestyle="--", alpha=0.8, zorder=1,
                       label=f"霧警戒ライン（{threshold_pct:.0f}%・地点ごとに最適化）")
        axp.set_ylim(0, 100)
        axp.set_yticks([0, 50, 100])
        axp.grid(True, axis="y", alpha=0.25)

        # ーーー 下段：この1地点だけの現象コード（実測＋予測、①②と同じ配色） ーーー
        if len(x_hist):
            slash_mask = hist_codes == 0
            if slash_mask.any():
                ax2.vlines(x_hist[slash_mask], -0.4, 0.4, color=SLASH_COLOR, linewidth=2.2, zorder=2)
            for code in range(1, 11):
                code_mask = hist_codes == code
                if code_mask.any():
                    ax2.vlines(x_hist[code_mask], -0.45, 0.45, color=PHENOM_COLORS[code],
                               linewidth=3.0, zorder=3)
        for code in range(0, 11):
            mask = pred_codes == code
            if mask.any():
                color = SLASH_COLOR if code == 0 else PHENOM_COLORS[code]
                ax2.vlines(x_fore[mask], -0.45, 0.45, color=color, linewidth=3.0, zorder=3)
        ax2.set_yticks([])
        ax2.set_ylim(-0.6, 0.6)
        ax2.axhline(0, color="#ececec", linewidth=0.5, zorder=1)

        span = (seg_end - seg_start).total_seconds() / 86400.0
        for ax in (ax1, axp, ax2):
            ax.set_xlim(seg_start, seg_end)
        _apply_date_axis(ax1, span)
        _apply_date_axis(axp, span)
        _apply_date_axis(ax2, span)

        # 左端のパネルにだけ軸ラベルを、右端にだけ湿度軸・凡例を出す
        if j == 0:
            ax1.set_ylabel("気温・露点温度（℃）", fontsize=10)
            axp.set_ylabel("霧の確率\n（％）", fontsize=9)
            ax2.set_ylabel(loc_name, fontsize=9)
            ax2.set_xlabel("実測（年/月/日）" if len(segments) > 1 else "日時（年/月/日）")
        else:
            ax2.set_xlabel("予報（年/月/日）")
        if is_last:
            ax1r.set_ylabel("相対湿度（％）", fontsize=10)
        else:
            ax1r.set_yticklabels([])

        if is_forecast_panel:
            # 凡例と重ならないよう、注記はパネルの見出しとして出す
            note = "ここから先は予測（背景グレー・点線）" if len(segments) == 1 else "予測ゾーン（背景グレー・点線）"
            ax1.set_title(note, loc="left", fontsize=9, color="dimgray")
        elif len(segments) > 1:
            ax1.set_title("実測", loc="left", fontsize=9, color="dimgray")

    # 凡例は右端のパネルにまとめて出す（各パネルに出すと図が読みにくくなるため）。
    # 上段はグラフ内に置くと線と重なって読みにくくなるため、パネルの外（上）に横並びで出す。
    ax1_last, axp_last, ax2_last = axes[0][-1], axes[1][-1], axes[2][-1]
    h1, lbl1 = axes[0][0].get_legend_handles_labels()
    h2, lbl2 = humid_axes[0].get_legend_handles_labels()
    axes[0][0].legend(h1 + h2, lbl1 + lbl2, bbox_to_anchor=(0, 1.02, 1, 0.12), loc="lower left",
                      ncol=3, mode="expand", borderaxespad=0., fontsize=7.5)
    if probs is not None:
        axp_last.legend(loc="upper left", fontsize=7.5)
    ax2_last.legend(handles=_phenom_legend_handles(), bbox_to_anchor=(1.08, 2.6), loc="upper left",
                    fontsize=7, borderaxespad=0., title="現象コード", title_fontsize=8)

    if len(segments) > 1:
        for row in range(3):
            left, right = axes[row][0], axes[row][1]
            left.spines["right"].set_visible(False)
            right.spines["left"].set_visible(False)
            right.tick_params(axis="y", left=False)
            _draw_axis_break_marks(left, right)
        gap_note = f"（実測の最終記録から予報開始まで{gap_days:.0f}日の空白があるため、横軸を分割しています）"
    else:
        gap_note = ""

    fig.suptitle(
        f"【{location_name}　{loc_name}】直近{history_days}日間の実測 ＋ 今後{n_days}日間の霧予測{gap_note}",
        fontsize=12, y=0.98,
    )
    fig.autofmt_xdate()

    out_path = os.path.join(
        out_dir,
        f"{location_name}_④{loc_name}_{n_days}日間予測_{fc_start.strftime('%Y%m%d')}.png",
    )
    fig.savefig(out_path, dpi=SAVE_DPI, bbox_inches="tight")
    plt.close(fig)
    return out_path


def plot_all_location_summary(pred_df, models, location_mapping, location_name, out_dir, prob_df=None):
    """全地点の予測をまとめる。

    上段: 日ごとに『霧(コード1〜6)が予測された地点数』の棒グラフ（＋霧確率の全地点平均の折れ線）。
    下段: 地点×日で「その日の霧確率の最大」を色の濃さで示すヒートマップ。
          どの地点がいつ危ないかを一覧できるようにするための追加パネル
          （以前は上段の棒グラフだけで、地点ごとの違いが見えなかった）。
    """
    df = pred_df.copy()
    df["date"] = df["datetime"].dt.date
    fog_cols = [c for c in models.keys() if c in df.columns]
    n_total = len(fog_cols)
    loc_names = [location_mapping.get(c, c) for c in fog_cols]

    dates = sorted(df["date"].unique())
    n_days = len(dates)
    counts, mean_probs = [], []
    prob_dates = prob_df["datetime"].dt.date if prob_df is not None else None
    for d in dates:
        group = df[df["date"] == d]
        counts.append(sum(1 for col in fog_cols if group[col].astype(int).isin(SUMMARY_FOG_CODES).any()))
        if prob_df is not None and fog_cols:
            mask = prob_dates == d
            mean_probs.append(float(prob_df.loc[mask, fog_cols].to_numpy().mean() * 100.0) if mask.any()
                              else np.nan)
        else:
            mean_probs.append(np.nan)

    # 地点×日の「その日の霧確率の最大」を格子状に並べる
    heat = np.full((n_total, n_days), np.nan)
    if prob_df is not None:
        for i, col in enumerate(fog_cols):
            for j, d in enumerate(dates):
                mask = prob_dates == d
                if mask.any():
                    heat[i, j] = float(prob_df.loc[mask, col].max() * 100.0)

    x = np.arange(n_days)
    fig_w = max(10, n_days * 0.8)
    fig_h = 3.0 + max(2.0, n_total * 0.22)
    fig, (ax_bar, ax_heat) = plt.subplots(
        2, 1, figsize=(fig_w, fig_h), sharex=True,
        gridspec_kw={"height_ratios": [3.0, max(2.0, n_total * 0.22)], "hspace": 0.08},
    )

    bars = ax_bar.bar(x, counts, color="#3498db", width=0.6, label="霧が予測された地点数")
    for bx, c in zip(x, counts):
        ax_bar.text(bx, c + 0.3, str(c), ha="center", va="bottom", fontsize=9)
    ax_bar.set_ylim(0, n_total + 1)
    ax_bar.yaxis.set_major_locator(mticker.MaxNLocator(integer=True))  # 地点数なので目盛りは整数だけ
    ax_bar.set_ylabel(f"霧が予測された\n地点数（全{n_total}中）", fontsize=9)
    ax_bar.set_title(f"【{location_name}】日ごとの霧予測サマリー（今後{n_days}日間）", fontsize=13)
    ax_bar.grid(True, axis="y", alpha=0.3)

    handles, labels = ax_bar.get_legend_handles_labels()
    if prob_df is not None and not all(np.isnan(mean_probs)):
        axr = ax_bar.twinx()
        line, = axr.plot(x, mean_probs, color="#c0392b", marker="o", markersize=4, linewidth=1.4,
                         label="霧確率の平均（全地点・全時刻）")
        axr.set_ylim(0, 100)
        axr.set_ylabel("霧確率\nの平均(%)", fontsize=9)
        handles.append(line)
        labels.append(line.get_label())
    ax_bar.legend(handles, labels, loc="upper left", fontsize=8)
    ax_bar.grid(True, axis="x", alpha=0.25, linewidth=0.6)

    # ーーー 下段：地点×日の霧確率ヒートマップ ーーー
    im = ax_heat.imshow(heat, aspect="auto", cmap="YlOrRd", vmin=0, vmax=100, interpolation="nearest")
    ax_heat.set_yticks(range(n_total))
    ax_heat.set_yticklabels(loc_names, fontsize=8)
    ax_heat.set_xticks(x)
    ax_heat.set_xticklabels([d.strftime("%m/%d") for d in dates], fontsize=8, rotation=45, ha="right")
    ax_heat.set_xlabel("日付")
    ax_heat.set_ylabel("地点", fontsize=9)
    for i in range(n_total):
        for j in range(n_days):
            v = heat[i, j]
            if not np.isnan(v) and v >= 50:
                ax_heat.text(j, i, f"{v:.0f}", ha="center", va="center", fontsize=6,
                             color="white" if v >= 70 else "black")
    cbar = fig.colorbar(im, ax=[ax_bar, ax_heat], fraction=0.02, pad=0.01)
    cbar.set_label("霧確率の最大（％）", fontsize=8)

    stamp = dates[0].strftime("%Y%m%d") if dates else "na"
    out_path = os.path.join(out_dir, f"{location_name}_⑤日別霧予測サマリー_{stamp}.png")
    fig.savefig(out_path, dpi=SAVE_DPI, bbox_inches="tight")
    plt.close(fig)
    return out_path


def _phenom_name(code: int) -> str:
    return "現象なし" if code == 0 else PHENOM_LABELS.get(code, f"コード{code}")


def export_prediction_csv(pred_df, prob_df, models, location_mapping, location_name, out_dir):
    """予測結果（地点ごとの予測コードと霧確率）をCSVに書き出す。

    既定は「縦持ち」形式（日時・地点・予測コード・現象名・霧確率・判定を1行ずつ並べる）。
    地点数が多いと横持ち（従来形式、地点ごとに列が並ぶ）はExcelで延々と右にスクロールする
    ことになるため、確認しやすい縦持ちを既定にした。横持ちも別ファイルとして両方出力する
    （ピボット表を作りたい場合や、従来の使い方に合わせたい場合向け）。

    「判定」列は、地点ごとに検証期間で最適化した霧確率のしきい値（`models[col]["fog_threshold"]`、
    既定50%ではなく地点ごとに変わる）を超えたら「霧」とする。
    """
    long_rows = []
    for col in models.keys():
        if col not in pred_df.columns:
            continue
        name = location_mapping.get(col, col)
        threshold = models[col].get("fog_threshold", DEFAULT_FOG_THRESHOLD)
        codes = pred_df[col].astype(int)
        probs = (prob_df[col] if prob_df is not None and col in prob_df.columns
                else pd.Series(np.nan, index=pred_df.index))
        part = pd.DataFrame({
            "日時": pred_df["datetime"].dt.strftime("%Y/%m/%d %H:%M"),
            "地点": name,
            "予測コード": codes,
            "現象名": codes.map(_phenom_name),
            "霧確率(%)": (probs * 100).round(1),
            "判定": np.where(probs >= threshold, "霧", "-"),
        })
        long_rows.append(part)

    stamp = pred_df["datetime"].min().strftime("%Y%m%d")
    long_out = pd.concat(long_rows, ignore_index=True) if long_rows else pd.DataFrame()
    out_path = os.path.join(out_dir, f"{location_name}_⑥予測結果_{stamp}.csv")
    long_out.to_csv(out_path, index=False, encoding="utf-8-sig")  # Excelで開けるようBOM付き

    # 従来形式（横持ち）も参考として出力する
    wide = pd.DataFrame({"日時": pred_df["datetime"].dt.strftime("%Y/%m/%d %H:%M")})
    for col in models.keys():
        if col not in pred_df.columns:
            continue
        name = location_mapping.get(col, col)
        wide[f"{name}_予測コード"] = pred_df[col].astype(int)
        if prob_df is not None and col in prob_df.columns:
            wide[f"{name}_霧確率(%)"] = (prob_df[col] * 100).round(1)
    wide_path = os.path.join(out_dir, f"{location_name}_⑥予測結果_横持ち_{stamp}.csv")
    wide.to_csv(wide_path, index=False, encoding="utf-8-sig")

    return out_path, wide_path


def load_or_train_models(main_df, phenom_df, phenom_cols, location_mapping, model, cache_path=None):
    """`cache_path` に保存済みモデルがあれば読み込み、無ければ学習してから保存する。

    グラフだけ作り直したいときに、毎回32地点分の学習をやり直さずに済ませるための機能。
    """
    if cache_path and os.path.isfile(cache_path):
        try:
            models = _joblib_load(cache_path)
            print(f"\n■ 学習済みモデルを再利用します: {cache_path}")
            print("  （学習をやり直したい場合は、このファイルを削除するか --model-cache に別のパスを指定してください）")
            return models
        except Exception as e:
            print(f"\n【警告】保存済みモデル（{cache_path}）の読み込みに失敗したため、学習をやり直します: "
                  f"{type(e).__name__}: {e}")

    models = train_location_models(main_df, phenom_df, phenom_cols, location_mapping, model=model)
    if cache_path and models:
        try:
            os.makedirs(os.path.dirname(os.path.abspath(cache_path)) or ".", exist_ok=True)
            _joblib_dump(models, cache_path)
            print(f"\n学習済みモデルを保存しました: {cache_path}"
                  "（次回は --model-cache で同じファイルを指定すると学習を省略できます）")
        except Exception as e:
            print(f"\n【警告】学習済みモデルの保存に失敗しました: {type(e).__name__}: {e}")
    return models


def _fetch_per_location_forecasts(coords_file, phenom_cols, location_mapping, forecast_days, forecast_raw):
    """`--location-coords` で指定されたCSVから、地点ごとの予報をまとめて取得する。

    只見川流域の32地点は数十kmにわたって散らばっているが、既定では全地点が
    只見町付近1地点分の予報を共用している。地点ごとの緯度経度が分かる場合は、
    このCSVを用意すると地点ごとに実際に近い予報値で予測できる。

    戻り値: {列文字: 予報DataFrame}（CSVに無い地点は空＝呼び出し側で既定の予報にフォールバックする）
    """
    try:
        coords_by_name = load_location_coordinates(coords_file)
    except (OSError, ValueError) as e:
        print(f"\n【警告】地点座標ファイルの読み込みに失敗しました（{coords_file}）: {e}")
        print("　→ 全地点で既定の1地点分の予報を使います。")
        return {}

    matched = {col: coords_by_name[location_mapping[col]]
              for col in phenom_cols
              if location_mapping.get(col) in coords_by_name}
    if not matched:
        print(f"\n【警告】{coords_file} の地点名が観測データの地点名と一致しませんでした。"
              "全地点で既定の1地点分の予報を使います。")
        return {}

    print(f"\n■ 地点ごとの座標ファイルを使用: {coords_file}"
          f"（{len(matched)}/{len(phenom_cols)}地点で個別の予報を取得）")
    name_to_coords = {location_mapping[col]: latlon for col, latlon in matched.items()}
    try:
        forecasts_by_name = fetch_forecast_multi(name_to_coords, forecast_days=forecast_days)
    except (requests.exceptions.RequestException, ValueError) as e:
        print(f"  （地点ごとの予報取得に失敗したため、既定の1地点分の予報を使います: {e}）")
        return {}

    # pred_df/prob_dfは共通のdatetime列（forecast_rawのもの）を軸にするため、
    # 地点ごとの予報が万一それと異なる時刻構成になっていたら、その地点だけ既定に戻す
    # （欠測補間の結果、まれに行数がずれることがあるため）。
    default_times = set(forecast_raw["datetime"])
    result, mismatched = {}, []
    for col in matched:
        name = location_mapping[col]
        df = forecasts_by_name.get(name)
        if df is None:
            continue
        if set(df["datetime"]) != default_times:
            mismatched.append(name)
            continue
        result[col] = df
    if mismatched:
        print(f"  （{len(mismatched)}地点は予報の時刻構成が既定と異なったため、既定の予報を使います: "
              f"{', '.join(mismatched)}）")
    return result


def run_fog_prediction_addon(main_df, phenom_df, phenom_cols, location_name, out_dir, location_mapping,
                             lat: float = None, lon: float = None, forecast_days: int = None,
                             history_days: int = 5, model=DEFAULT_MODEL_KIND, model_cache=None,
                             location_coords_file=None):
    """モデル学習 → 予報取得 → 地点ごとの予測グラフ・サマリー・CSV出力までを行う。"""
    lat = FORECAST_LAT if lat is None else lat
    lon = FORECAST_LON if lon is None else lon
    forecast_days = FORECAST_DAYS if forecast_days is None else forecast_days

    report_phenomena_quality(phenom_df, phenom_cols, location_mapping)

    models = load_or_train_models(main_df, phenom_df, phenom_cols, location_mapping, model, model_cache)
    if not models:
        print("\n【予測モデル】1地点も学習できなかったため、予測グラフの生成をスキップしました。")
        return []

    print("\n" + "=" * 66)
    print(f"■ 今後{forecast_days}日間の気象予報から地点ごとの現象コード予測グラフを生成")
    print("=" * 66)
    print(f"  取得元URL: {OPEN_METEO_URL}")
    print(f"  取得地点 : 緯度{lat}, 経度{lon}（--lat / --lon で変更できます）")
    try:
        forecast_raw = fetch_forecast_range(lat=lat, lon=lon, forecast_days=forecast_days)
    except (requests.exceptions.RequestException, ValueError) as e:
        print(f"\n(予報取得に失敗しました: {e})")
        print("→ インターネットに出られる環境で実行してください。")
        return []

    print(f"  取得件数 : {len(forecast_raw)}件")
    print(f"  取得期間 : {forecast_raw['datetime'].min()} 〜 {forecast_raw['datetime'].max()}")
    print("  先頭3行:")
    print(forecast_raw.head(3).to_string(index=False))

    forecast_by_col = {}
    if location_coords_file:
        forecast_by_col = _fetch_per_location_forecasts(
            location_coords_file, phenom_cols, location_mapping, forecast_days, forecast_raw)

    pred_df, prob_df = build_location_forecast_codes(models, forecast_raw, forecast_by_col)

    generated_paths = []
    location_summaries = []
    target_cols = [c for c in phenom_cols if c in models]
    progress = _ProgressBar(len(target_cols), prefix="④ 地点別予測グラフ")
    gap_warnings = []
    for col in target_cols:
        loc_name = location_mapping.get(col, col)
        progress.update(1, suffix=loc_name)
        col_forecast = forecast_by_col.get(col, forecast_raw)

        last_observed = get_last_observed_datetime(phenom_df, col)
        gap_days = (col_forecast["datetime"].min() - last_observed).total_seconds() / 86400
        if gap_days > 3:
            gap_warnings.append(
                f"{loc_name}: 最後の観測記録（{last_observed.strftime('%Y/%m/%d')}）から"
                f"予報開始（{col_forecast['datetime'].min().strftime('%Y/%m/%d')}）まで"
                f"{gap_days:.1f}日の空白があります。")

        out_path = plot_single_location_forecast(
            main_df, phenom_df, col, location_mapping, col_forecast, pred_df[col],
            location_name, out_dir, history_days=history_days, prob_series=prob_df[col],
            fog_threshold=models[col].get("fog_threshold", DEFAULT_FOG_THRESHOLD),
        )
        generated_paths.append(out_path)

        counts = pred_df[col].astype(int).value_counts().sort_index()
        breakdown = ", ".join(f"{('/' if c == 0 else c)}:{n}時間" for c, n in counts.items())
        max_prob = float(prob_df[col].max()) * 100
        location_summaries.append((loc_name, os.path.basename(out_path), breakdown, max_prob))
    progress.close()

    if gap_warnings:
        print("\n【注意】実測の最終記録から予報開始まで間が空いている地点があります:")
        for w in gap_warnings:
            print(f"  {w}")

    print()
    for loc_name, fname, breakdown, max_prob in location_summaries:
        print(f"  [{loc_name}] {fname}")
        print(f"      予測内訳: {breakdown}　／　霧確率の最大: {max_prob:.0f}%")

    print(f"\n{len(generated_paths)}地点分の予測グラフを生成しました（{out_dir} 内）。")

    summary_path = plot_all_location_summary(pred_df, models, location_mapping, location_name,
                                             out_dir, prob_df=prob_df)
    print(f"全地点サマリーグラフ: {os.path.basename(summary_path)}")
    csv_path, csv_wide_path = export_prediction_csv(
        pred_df, prob_df, models, location_mapping, location_name, out_dir)
    print(f"予測結果CSV        : {os.path.basename(csv_path)}"
          f"（横持ち版: {os.path.basename(csv_wide_path)}）")
    return generated_paths + [summary_path, csv_path, csv_wide_path]


# ===========================================================================
# 7. メイン実行部および Google Colab 支援機能
# ===========================================================================

DEFAULT_INPUT_FILE = "気象データ.xlsx"
DEFAULT_OUTPUT_DIR = "./output_graphs"
DATA_FILE_EXTENSIONS = (".xlsx", ".xlsm", ".xls", ".csv")


def _in_notebook():
    return "google.colab" in sys.modules or "ipykernel" in sys.modules


def _try_colab_upload():
    if "google.colab" not in sys.modules:
        return None
    try:
        from google.colab import files as colab_files
    except ImportError:
        return None
    print("【確認】入力ファイルが指定のパスに見つかりません。")
    print("アップロード画面を表示しますので、解析したい気象データのファイルを選択してください…")
    uploaded = colab_files.upload()
    for name in uploaded.keys():
        if name.lower().endswith(DATA_FILE_EXTENSIONS):
            return os.path.abspath(name)
    return None


def _resolve_input_file(filepath):
    if filepath and os.path.isfile(filepath):
        return filepath
    if "google.colab" in sys.modules:
        uploaded_path = _try_colab_upload()
        if uploaded_path and os.path.isfile(uploaded_path):
            return uploaded_path
    if _in_notebook():
        try:
            entered = input("入力ファイルのパスを入力してください（Enterでキャンセル): ").strip()
            if entered and os.path.isfile(entered):
                return entered
        except Exception:
            pass
    return None


def _parse_args(argv=None):
    """コマンドライン引数を解釈する。

    Jupyter/Colabでは sys.argv にカーネルの引数（-f /.../kernel.json など）が
    入っており、それを出力フォルダと誤認する不具合があったため、
    ノートブック環境では引数を読まずに既定値を使う。
    """
    if argv is None:
        argv = [] if _in_notebook() else sys.argv[1:]

    parser = argparse.ArgumentParser(
        description="只見川 川霧 観測・予測プログラム",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("input", nargs="?", default=None,
                        help="入力ファイル（.xlsx / .xlsm / .csv）")
    parser.add_argument("output", nargs="?", default=None, help="出力フォルダ")
    parser.add_argument("--sheet", default=None, help="Excelのシート名（省略時は先頭シート）")
    parser.add_argument("--layout", choices=["auto", "fixed"], default="auto",
                        help="列の決め方（auto=見出しから自動検出 / fixed=従来の固定列）")
    parser.add_argument("--lat", type=float, default=FORECAST_LAT, help="予報を取得する地点の緯度")
    parser.add_argument("--lon", type=float, default=FORECAST_LON, help="予報を取得する地点の経度")
    parser.add_argument("--forecast-days", type=int, default=FORECAST_DAYS,
                        help="予報を取得する日数（Open-Meteoの上限は16）")
    parser.add_argument("--history-days", type=int, default=5,
                        help="④のグラフに表示する実測データの日数")
    parser.add_argument("--model", choices=list(MODEL_KINDS), default=DEFAULT_MODEL_KIND,
                        help="学習モデル（hybrid=RandomForest+XGBoostの加重平均 / "
                             "rf=RandomForest単独＝従来と同じ速さ / xgb=XGBoost単独）")
    parser.add_argument("--model-cache", default=None,
                        help="学習済みモデルの保存/再利用先ファイル（.joblib）。"
                             "指定すると2回目以降は学習を省略できる")
    parser.add_argument("--location-coords", default=None,
                        help="地点名,緯度,経度の列を持つCSV。指定すると地点ごとに予報を取得する"
                             "（未指定なら全地点で --lat/--lon の1地点分の予報を共用）")
    parser.add_argument("--no-forecast", action="store_true", help="霧予測（④⑤⑥）を行わない")
    parser.add_argument("--no-monthly", action="store_true", help="月別グラフ（①②）を作らない")
    parser.add_argument("--no-log", action="store_true", help="実行ログのファイル保存をしない")
    parser.add_argument("--font", default=None,
                        help="グラフに使う日本語フォント名（例: IPAexGothic）")
    parser.add_argument("--check-font", action="store_true",
                        help="日本語フォントの状態を確認し、確認用の画像だけを出力して終了する")
    parser.add_argument("--zip", action="store_true",
                        help="出力フォルダをZIPにまとめる（Colabから持ち帰るとき用）")
    args = parser.parse_args(argv)

    # 「フォルダ 入力ファイル」の順で渡された場合も受け付ける（従来の柔軟な指定に合わせる）
    def looks_like_data_file(p):
        return bool(p) and p.lower().endswith(DATA_FILE_EXTENSIONS)

    if not looks_like_data_file(args.input) and looks_like_data_file(args.output):
        args.input, args.output = args.output, args.input
    elif args.input and not looks_like_data_file(args.input) and args.output is None:
        # 引数が1つだけで、それがデータファイルに見えない場合は出力フォルダとして扱う
        args.input, args.output = None, args.input
    if args.input is None:
        args.input = DEFAULT_INPUT_FILE
    if args.output is None:
        args.output = DEFAULT_OUTPUT_DIR
    return args


def plot_font_check(out_dir):
    """日本語が正しく描けるか確認するための小さな画像を出力する。

    グラフ本体を作る前に「文字化けしていないか」だけを数秒で確かめられるようにする。
    """
    supported, used_name, used_path = verify_japanese_font()
    fig, ax = plt.subplots(figsize=(9, 3.2))
    ax.axis("off")
    ax.text(0.5, 0.72, "日本語フォント確認：只見川の川霧 予測グラフ", ha="center", fontsize=16)
    ax.text(0.5, 0.45, "気温(℃)・露点温度(℃)・相対湿度(％)・風速(m/s)・降水量(mm)",
            ha="center", fontsize=12)
    ax.text(0.5, 0.20, f"使用フォント: {used_name}", ha="center", fontsize=10, color="dimgray")
    out_path = os.path.join(out_dir, "font_check.png")
    fig.savefig(out_path, dpi=SAVE_DPI, bbox_inches="tight")
    plt.close(fig)

    print(f"確認用の画像を出力しました: {out_path}")
    if supported:
        print("この画像の文字が読めれば、グラフの日本語は問題ありません。")
    else:
        print("上の警告のとおり、日本語が □ になる可能性が高い状態です。")
    return out_path


def display_in_notebook(paths, max_images: int = 8, width: int = 1100):
    """Colabのセル内に、生成したグラフをそのまま表示する。

    出力フォルダを開かなくても結果が確認できるようにするための表示処理。
    枚数が多いときは max_images 枚だけ表示し、残りはファイル名だけ知らせる。
    """
    if not _in_notebook():
        return []
    try:
        from IPython.display import Image, display
    except ImportError:
        return []

    images = [p for p in paths if str(p).lower().endswith(".png")]
    shown = images[:max_images]
    for path in shown:
        print(os.path.basename(path))
        display(Image(filename=path, width=width))
    if len(images) > len(shown):
        print(f"（ほか{len(images) - len(shown)}枚は出力フォルダに保存されています）")
    return shown


def zip_outputs(out_dir, download: bool = False):
    """出力フォルダをZIPにまとめる（Colabから手元に持ち帰るとき用）。"""
    base = os.path.abspath(out_dir.rstrip("/"))
    zip_path = shutil.make_archive(base, "zip", out_dir)
    print(f"ZIPにまとめました: {zip_path}")
    if download and "google.colab" in sys.modules:
        try:
            from google.colab import files as colab_files
            colab_files.download(zip_path)
        except Exception as e:
            print(f"（自動ダウンロードに失敗しました: {e}　左のファイル一覧から手動で保存してください）")
    return zip_path


class _TeeOutput:
    """print()の出力を、画面と実行ログファイルの両方に同時に書き出す。"""

    def __init__(self, *streams):
        self.streams = streams

    def write(self, s):
        for st in self.streams:
            st.write(s)

    def flush(self):
        for st in self.streams:
            st.flush()


def run(input_file=None, output_dir=None, sheet=None, layout="auto",
        lat=None, lon=None, forecast_days=None, history_days=5,
        forecast=True, monthly=True, font=None, show=None,
        zip_output=False, download=False, model=DEFAULT_MODEL_KIND,
        model_cache=None, location_coords_file=None, save_log=True):
    """Colabのセルから直接呼べるエントリーポイント。

        from weather_visualizer import run
        run("/content/drive/MyDrive/只見_気象データ.xlsx")

    引数はコマンドラインのオプションと同じ意味。show=True でグラフをセル内に表示、
    zip_output=True で出力フォルダをZIPにまとめる（download=Trueでダウンロードまで）。
    model="hybrid"（既定）はRandomForestとXGBoostのハイブリッド、"rf" / "xgb" で
    片方だけを使う（"rf" は従来と同じ速さ）。
    model_cache に .joblib のパスを指定すると、学習済みモデルを保存/再利用できる
    （2回目以降はモデル学習を省略してグラフだけ作り直せる）。
    location_coords_file に「地点名,緯度,経度」のCSVを指定すると、地点ごとに
    予報を取得する（未指定なら全地点で lat/lon の1地点分の予報を共用する）。
    save_log=False にすると実行ログ（実行ログ_日時.txt）をファイルに保存しない。
    """
    if font:
        setup_japanese_font(font_name=font)

    out_dir = output_dir or DEFAULT_OUTPUT_DIR
    filepath = _resolve_input_file(input_file or DEFAULT_INPUT_FILE)
    if filepath is None:
        raise FileNotFoundError(
            f"入力ファイルが見つかりません: {input_file or DEFAULT_INPUT_FILE}\n"
            "　対応形式: .xlsx / .xlsm / .csv\n"
            "　Colabでは左のファイル一覧にアップロードするか、Googleドライブをマウントして"
            "そのパスを指定してください。"
        )

    os.makedirs(out_dir, exist_ok=True)

    log_path = None
    original_stdout = sys.stdout
    log_file = None
    if save_log:
        log_path = os.path.join(out_dir, f"実行ログ_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
        log_file = open(log_path, "w", encoding="utf-8")
        sys.stdout = _TeeOutput(original_stdout, log_file)

    try:
        print(f"ファイルを読み込み中: {filepath}")
        main_df, phenom_df, phenom_cols, location_mapping = load_weather_data(
            filepath, sheet_name=sheet, layout=layout)

        base = os.path.splitext(os.path.basename(filepath))[0]
        location_name = base.split("_")[0].split(" ")[0] if ("_" in base or " " in base) else base

        generated = []
        if monthly:
            print(f"\n月ごと・全地点統合グラフの生成を開始します（{len(phenom_cols)}地点）...")
            made = plot_combo_by_month(main_df, phenom_df, phenom_cols, location_mapping,
                                       location_name, out_dir)
            generated += made
            print(f"月別グラフを{len(made)}枚生成しました。出力先: {os.path.abspath(out_dir)}")
        else:
            print("\n月別グラフ（①②）の生成はスキップしました。")

        if forecast:
            generated += run_fog_prediction_addon(
                main_df, phenom_df, phenom_cols, location_name, out_dir, location_mapping,
                lat=lat, lon=lon, forecast_days=forecast_days, history_days=history_days,
                model=model, model_cache=model_cache, location_coords_file=location_coords_file,
            )
        else:
            print("\n霧予測（④⑤⑥）の生成はスキップしました。")

        print(f"\n完了しました。出力先フォルダ: {os.path.abspath(out_dir)}")
        if log_path:
            print(f"実行ログを保存しました: {log_path}")
    finally:
        sys.stdout = original_stdout
        if log_file:
            log_file.close()

    if show is None:
        show = _in_notebook()  # ノートブックなら既定で表示する
    if show:
        # 予測グラフ（④⑤）を優先して表示する
        priority = [p for p in generated if "④" in os.path.basename(p) or "⑤" in os.path.basename(p)]
        display_in_notebook(priority + [p for p in generated if p not in priority])
    if zip_output:
        zip_outputs(out_dir, download=download)
    return generated


def main(argv=None):
    args = _parse_args(argv)
    if args.font:
        setup_japanese_font(font_name=args.font)
    elif args.check_font and not verify_japanese_font()[0]:
        # 確認モードで日本語が描けない状態なら、もう一度導入を試みる
        setup_japanese_font()

    if args.check_font:
        os.makedirs(args.output, exist_ok=True)
        plot_font_check(args.output)
        return 0

    try:
        run(input_file=args.input, output_dir=args.output, sheet=args.sheet, layout=args.layout,
            lat=args.lat, lon=args.lon, forecast_days=args.forecast_days,
            history_days=args.history_days, forecast=not args.no_forecast,
            monthly=not args.no_monthly, zip_output=args.zip, model=args.model,
            model_cache=args.model_cache, location_coords_file=args.location_coords,
            save_log=not args.no_log)
    except FileNotFoundError as e:
        print(f"【エラー】{e}")
        return 1
    return 0


# ===========================================================================
# グラフの見方（メモ）
# ---------------------------------------------------------------------------
# ・日付ラベルの位置はその日の0時です。
# ・現象コードの色一覧
#      /   薄い黄色   「現象なし」
#      1   薄い水色   「薄い川霧」
#      2   水色       「川霧」
#      3   濃い水色   「濃い川霧」
#      4   薄い灰色   「薄い全体霧」
#      5   灰色       「全体霧」
#      6   紺色       「全体濃い霧」
#      7   オレンジ色 「薄い層雲」
#      8   赤茶色     「濃い層雲」
#      9   薄い緑色   「霧雨」
#      10  緑色       「雨」
# ===========================================================================


if __name__ == "__main__":
    sys.exit(main())
